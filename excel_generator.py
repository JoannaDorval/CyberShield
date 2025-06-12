"""
Excel Report Generator for Enhanced TARA Analysis
Generates comprehensive Excel reports with cybersecurity assessment data
"""

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import logging
from typing import Dict, List, Any, Optional
from datetime import datetime

class TaraExcelGenerator:
    """Generate comprehensive TARA Excel reports with multiple worksheets"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        self.workbook = None
        self.worksheets = {}
    
    def generate_excel_report(self, analysis_data: Dict[str, Any], 
                            input_type: str, 
                            cross_ref_source: str,
                            embed_assessment: Optional[Dict] = None) -> str:
        """
        Generate complete Excel report with all analysis data
        
        Args:
            analysis_data: Complete analysis results
            input_type: 'threat_model', 'block_diagram', or 'both'
            cross_ref_source: 'mitre_attack', 'mitre_embed', or 'both'
            embed_assessment: MITRE EMBED assessment data if applicable
        
        Returns:
            str: Path to generated Excel file
        """
        try:
            # Create new workbook
            self.workbook = Workbook()
            
            # Remove default sheet
            default_sheet = self.workbook.active
            if default_sheet:
                self.workbook.remove(default_sheet)
            
            # Create single consolidated worksheet
            self._create_consolidated_worksheet(analysis_data, input_type, cross_ref_source, embed_assessment)
            
            # Generate filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"TARA_Excel_Report_{timestamp}.xlsx"
            
            # Create uploads directory if it doesn't exist
            output_dir = 'uploads'
            os.makedirs(output_dir, exist_ok=True)
            
            filepath = os.path.join(output_dir, filename)
            
            # Save workbook
            self.workbook.save(filepath)
            self.logger.info(f"Excel report generated: {filepath}")
            
            return filepath
            
        except Exception as e:
            self.logger.error(f"Failed to generate Excel report: {e}")
            raise
    
    def _create_consolidated_worksheet(self, analysis_data: Dict[str, Any], input_type: str, cross_ref_source: str, embed_assessment: Optional[Dict] = None):
        """Create single consolidated worksheet with all sections arranged horizontally"""
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        # Create the main worksheet
        ws = self.workbook.create_sheet('TARA Analysis Report')
        self.workbook.active = ws
        
        # Define color schemes for each section
        colors = {
            'assets': 'E7E6E6',
            'damage_scenario': 'E2EFDA', 
            'impact_analysis': 'FFF2CC',
            'cybersecurity_controls': 'F2F2F2',
            'threat_scenarios': 'DEEAF6',
            'attack_path': 'FCE4D6',
            'attack_feasibility_assessment': 'EDEDED',
            'cybersecurity_goals_and_claims': 'E1D5E7'
        }
        
        # Define section headers and their columns
        sections = {
            'Assets': ['Asset ID', 'Assets', 'Security Property Loss'],
            'Damage Scenario': ['Stakeholder', 'Damage Scenario Description'],
            'Impact Analysis': ['Impact Category', 'Impact Level', 'Impact Description'],
            'Cybersecurity Controls': ['Control ID', 'Control Name', 'Control Description'],
            'Threat Scenarios': ['Threat ID', 'Threat Name', 'STRIDE Category'],
            'Attack Path': ['Attack Vector', 'Attack Steps', 'Prerequisites'],
            'Attack Feasibility Assessment': ['Feasibility Factor', 'Rating', 'Justification'],
            'Cybersecurity Goals and Claims': ['Goal ID', 'Security Goal', 'Claim Statement']
        }
        
        # Start position
        current_col = 1
        
        # Create each section
        for section_name, columns in sections.items():
            color_key = section_name.lower().replace(' ', '_')
            fill_color = colors.get(color_key, 'E7E6E6')  # Default gray
            
            # Create section header
            self._create_section_header(ws, section_name, current_col, len(columns), fill_color)
            
            # Create column headers
            for i, col_name in enumerate(columns):
                cell = ws.cell(row=2, column=current_col + i)
                cell.value = col_name
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                cell.font = Font(bold=True, size=10)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            # Fill section data
            self._fill_section_data(ws, section_name, analysis_data, current_col, fill_color, embed_assessment)
            
            # Move to next section
            current_col += len(columns) + 1  # Add gap between sections
        
        # Auto-fit column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_section_header(self, ws, section_name, start_col, num_cols, fill_color):
        """Create merged header for section"""
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        # Merge cells for header
        if num_cols > 1:
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=start_col + num_cols - 1)
        
        # Style the header
        header_cell = ws.cell(row=1, column=start_col)
        header_cell.value = section_name
        header_cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        header_cell.font = Font(bold=True, size=12)
        header_cell.alignment = Alignment(horizontal='center', vertical='center')
        header_cell.border = Border(
            left=Side(style='thick'),
            right=Side(style='thick'),
            top=Side(style='thick'),
            bottom=Side(style='thick')
        )
    
    def _fill_section_data(self, ws, section_name, analysis_data, start_col, fill_color, embed_assessment):
        """Fill data for each section"""
        from openpyxl.styles import PatternFill, Alignment, Border, Side
        
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        row = 3  # Start data from row 3
        
        if section_name == 'Assets':
            assets = analysis_data.get('assets', [])
            for i, asset in enumerate(assets):
                ws.cell(row=row + i, column=start_col).value = f"A{i+1:03d}"
                ws.cell(row=row + i, column=start_col + 1).value = asset.get('name', 'Unknown')
                ws.cell(row=row + i, column=start_col + 2).value = asset.get('criticality', 'Medium')
                
                for col in range(3):
                    cell = ws.cell(row=row + i, column=start_col + col)
                    cell.fill = fill
                    cell.border = border
                    cell.alignment = Alignment(vertical='center')
        
        elif section_name == 'Damage Scenario':
            threats = analysis_data.get('threats', [])
            stakeholders = ['Users', 'Organization', 'Partners', 'Regulators']
            
            scenario_row = row
            for threat in threats[:10]:  # Limit to first 10 threats
                for stakeholder in stakeholders:
                    if scenario_row - row < 20:  # Limit total scenarios
                        ws.cell(row=scenario_row, column=start_col).value = stakeholder
                        scenario_desc = self._generate_damage_scenario(threat, stakeholder)
                        ws.cell(row=scenario_row, column=start_col + 1).value = scenario_desc
                        
                        for col in range(2):
                            cell = ws.cell(row=scenario_row, column=start_col + col)
                            cell.fill = fill
                            cell.border = border
                            cell.alignment = Alignment(vertical='center', wrap_text=True)
                        
                        scenario_row += 1
        
        elif section_name == 'Impact Analysis':
            threats = analysis_data.get('threats', [])
            for i, threat in enumerate(threats[:15]):  # Limit to 15 threats
                impact = self._calculate_impact_analysis(threat)
                ws.cell(row=row + i, column=start_col).value = impact.get('category', 'Operational')
                ws.cell(row=row + i, column=start_col + 1).value = impact.get('level', 'Medium')
                ws.cell(row=row + i, column=start_col + 2).value = impact.get('description', 'Standard impact assessment')
                
                for col in range(3):
                    cell = ws.cell(row=row + i, column=start_col + col)
                    cell.fill = fill
                    cell.border = border
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
        
        elif section_name == 'Cybersecurity Controls':
            controls = analysis_data.get('mitigations', [])
            if embed_assessment:
                controls.extend(embed_assessment.get('controls', []))
            
            for i, control in enumerate(controls[:20]):  # Limit to 20 controls
                ws.cell(row=row + i, column=start_col).value = f"C{i+1:03d}"
                ws.cell(row=row + i, column=start_col + 1).value = control.get('name', control.get('title', 'Security Control'))
                ws.cell(row=row + i, column=start_col + 2).value = control.get('description', 'Standard security control')
                
                for col in range(3):
                    cell = ws.cell(row=row + i, column=start_col + col)
                    cell.fill = fill
                    cell.border = border
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
        
        elif section_name == 'Threat Scenarios':
            threats = analysis_data.get('threats', [])
            for i, threat in enumerate(threats[:15]):
                stride = self._analyze_stride_threat(threat)
                ws.cell(row=row + i, column=start_col).value = f"T{i+1:03d}"
                ws.cell(row=row + i, column=start_col + 1).value = threat.get('name', threat.get('title', 'Security Threat'))
                ws.cell(row=row + i, column=start_col + 2).value = stride.get('primary_category', 'Information Disclosure')
                
                for col in range(3):
                    cell = ws.cell(row=row + i, column=start_col + col)
                    cell.fill = fill
                    cell.border = border
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
        
        elif section_name == 'Attack Path':
            threats = analysis_data.get('threats', [])
            for i, threat in enumerate(threats[:12]):
                attack_path = self._generate_attack_path(threat, analysis_data)
                ws.cell(row=row + i, column=start_col).value = attack_path.get('vector', 'Network')
                ws.cell(row=row + i, column=start_col + 1).value = attack_path.get('steps', 'Multi-step attack sequence')
                ws.cell(row=row + i, column=start_col + 2).value = attack_path.get('prerequisites', 'Network access required')
                
                for col in range(3):
                    cell = ws.cell(row=row + i, column=start_col + col)
                    cell.fill = fill
                    cell.border = border
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
        
        elif section_name == 'Attack Feasibility Assessment':
            threats = analysis_data.get('threats', [])
            feasibility_factors = ['Technical Skill', 'Resources Required', 'Detection Likelihood']
            
            factor_row = row
            for threat in threats[:8]:  # Limit threats
                for factor in feasibility_factors:
                    if factor_row - row < 24:  # Limit total assessments
                        feasibility = self._assess_attack_feasibility(threat)
                        ws.cell(row=factor_row, column=start_col).value = factor
                        ws.cell(row=factor_row, column=start_col + 1).value = feasibility.get('rating', 'Medium')
                        ws.cell(row=factor_row, column=start_col + 2).value = feasibility.get('justification', 'Standard assessment')
                        
                        for col in range(3):
                            cell = ws.cell(row=factor_row, column=start_col + col)
                            cell.fill = fill
                            cell.border = border
                            cell.alignment = Alignment(vertical='center', wrap_text=True)
                        
                        factor_row += 1
        
        elif section_name == 'Cybersecurity Goals and Claims':
            threats = analysis_data.get('threats', [])
            mitigations = analysis_data.get('mitigations', [])
            
            for i, threat in enumerate(threats[:12]):
                goal = self._generate_cybersecurity_goal(threat, mitigations)
                ws.cell(row=row + i, column=start_col).value = f"G{i+1:03d}"
                ws.cell(row=row + i, column=start_col + 1).value = goal.get('goal', 'Ensure system security')
                ws.cell(row=row + i, column=start_col + 2).value = goal.get('claim', 'Security controls effectively mitigate threats')
                
                for col in range(3):
                    cell = ws.cell(row=row + i, column=start_col + col)
                    cell.fill = fill
                    cell.border = border
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
    
    def _create_assets_worksheet(self, analysis_data: Dict[str, Any]):
        """Create Assets worksheet"""
        ws = self.workbook.create_sheet("Assets")
        
        # Headers
        headers = ["Asset ID", "Asset", "Security Property Loss"]
        ws.append(headers)
        
        # Style headers
        self._style_headers(ws, len(headers))
        
        # Add asset data
        assets = analysis_data.get('assets', [])
        for i, asset in enumerate(assets, 1):
            row = [
                f"A{i:03d}",
                asset.get('name', 'Unknown Asset'),
                asset.get('security_property', 'Confidentiality, Integrity, Availability')
            ]
            ws.append(row)
        
        # Auto-adjust column widths
        self._adjust_column_widths(ws)
        
        self.worksheets['assets'] = ws
    
    def _create_damage_scenarios_worksheet(self, analysis_data: Dict[str, Any]):
        """Create Damage Scenarios worksheet"""
        ws = self.workbook.create_sheet("Damage Scenarios")
        
        headers = ["Stakeholder", "Damage Scenario Description"]
        ws.append(headers)
        self._style_headers(ws, len(headers))
        
        # Generate damage scenarios from threats
        threats = analysis_data.get('threats', [])
        stakeholders = ['End Users', 'Organization', 'Third Parties', 'Regulators']
        
        for threat in threats:
            for stakeholder in stakeholders:
                scenario = self._generate_damage_scenario(threat, stakeholder)
                if scenario:
                    ws.append([stakeholder, scenario])
        
        self._adjust_column_widths(ws)
        self.worksheets['damage_scenarios'] = ws
    
    def _create_impact_analysis_worksheet(self, analysis_data: Dict[str, Any]):
        """Create Impact Analysis worksheet"""
        ws = self.workbook.create_sheet("Impact Analysis")
        
        headers = [
            "Threat ID", "Safety", "Financial", "Operational", 
            "Privacy", "Impact Level", "Justification"
        ]
        ws.append(headers)
        self._style_headers(ws, len(headers))
        
        threats = analysis_data.get('threats', [])
        for threat in threats:
            impact_data = self._calculate_impact_analysis(threat)
            row = [
                threat.get('id', 'T-XXX'),
                impact_data['safety'],
                impact_data['financial'],
                impact_data['operational'],
                impact_data['privacy'],
                impact_data['overall_level'],
                impact_data['justification']
            ]
            ws.append(row)
        
        self._adjust_column_widths(ws)
        self.worksheets['impact_analysis'] = ws
    
    def _create_cybersecurity_controls_worksheet(self, analysis_data: Dict[str, Any], 
                                               embed_assessment: Optional[Dict] = None):
        """Create Cybersecurity Controls worksheet"""
        ws = self.workbook.create_sheet("Cybersecurity Controls")
        
        headers = ["ID", "Cybersecurity Control"]
        ws.append(headers)
        self._style_headers(ws, len(headers))
        
        # Add existing mitigations
        mitigations = analysis_data.get('mitigations', [])
        for mitigation in mitigations:
            row = [
                mitigation.get('id', 'C-XXX'),
                mitigation.get('description', mitigation.get('name', 'Unknown Control'))
            ]
            ws.append(row)
        
        # Add EMBED controls if available
        if embed_assessment and 'recommended_controls' in embed_assessment:
            for control in embed_assessment['recommended_controls']:
                row = [
                    control.get('id', 'EMBED-XXX'),
                    control.get('control', control.get('description', 'EMBED Control'))
                ]
                ws.append(row)
        
        self._adjust_column_widths(ws)
        self.worksheets['cybersecurity_controls'] = ws
    
    def _create_threat_scenarios_worksheet(self, analysis_data: Dict[str, Any]):
        """Create Threat Scenarios worksheet (STRIDE)"""
        ws = self.workbook.create_sheet("Threat Scenarios")
        
        headers = [
            "Threat ID", "Spoofing", "Tampering", "Repudiation", 
            "Information Disclosure", "Denial of Service", "Elevation of Privileges"
        ]
        ws.append(headers)
        self._style_headers(ws, len(headers))
        
        threats = analysis_data.get('threats', [])
        for threat in threats:
            stride_analysis = self._analyze_stride_threat(threat)
            row = [
                threat.get('id', 'T-XXX'),
                stride_analysis['spoofing'],
                stride_analysis['tampering'],
                stride_analysis['repudiation'],
                stride_analysis['information_disclosure'],
                stride_analysis['denial_of_service'],
                stride_analysis['elevation_of_privileges']
            ]
            ws.append(row)
        
        self._adjust_column_widths(ws)
        self.worksheets['threat_scenarios'] = ws
    
    def _create_attack_paths_worksheet(self, analysis_data: Dict[str, Any]):
        """Create Attack Paths worksheet"""
        ws = self.workbook.create_sheet("Attack Paths")
        
        headers = ["Threat ID", "Attack Path", "Entry Point", "Target Asset", "Attack Steps"]
        ws.append(headers)
        self._style_headers(ws, len(headers))
        
        threats = analysis_data.get('threats', [])
        for threat in threats:
            attack_path = self._generate_attack_path(threat, analysis_data)
            row = [
                threat.get('id', 'T-XXX'),
                attack_path.get('path_name', 'Standard Attack Path'),
                attack_path.get('entry_point', 'External Network'),
                attack_path.get('target', threat.get('target', 'Unknown')),
                attack_path.get('steps', 'Multi-step attack sequence')
            ]
            ws.append(row)
        
        self._adjust_column_widths(ws)
        self.worksheets['attack_paths'] = ws
    
    def _create_attack_feasibility_worksheet(self, analysis_data: Dict[str, Any]):
        """Create Attack Feasibility Assessment worksheet"""
        ws = self.workbook.create_sheet("Attack Feasibility")
        
        headers = [
            "Threat ID", "Elapsed Time", "Specialist Expertise", "Knowledge of Item", 
            "Window of Opportunity", "Equipment", "Attack Vector", 
            "Summary Attack Feasibility", "Risk Determination"
        ]
        ws.append(headers)
        self._style_headers(ws, len(headers))
        
        threats = analysis_data.get('threats', [])
        for threat in threats:
            feasibility = self._assess_attack_feasibility(threat)
            row = [
                threat.get('id', 'T-XXX'),
                feasibility['elapsed_time'],
                feasibility['specialist_expertise'],
                feasibility['knowledge_of_item'],
                feasibility['window_of_opportunity'],
                feasibility['equipment'],
                feasibility['attack_vector'],
                feasibility['summary_feasibility'],
                feasibility['risk_determination']
            ]
            ws.append(row)
        
        self._adjust_column_widths(ws)
        self.worksheets['attack_feasibility'] = ws
    
    def _create_risk_evaluation_worksheet(self, analysis_data: Dict[str, Any]):
        """Create Risk Evaluation worksheet"""
        ws = self.workbook.create_sheet("Risk Evaluation")
        
        headers = [
            "Threat ID", "Likelihood", "Impact", "Risk Level", 
            "Risk Score", "Risk Category", "Treatment Required"
        ]
        ws.append(headers)
        self._style_headers(ws, len(headers))
        
        threats = analysis_data.get('threats', [])
        for threat in threats:
            risk_eval = self._evaluate_risk(threat)
            row = [
                threat.get('id', 'T-XXX'),
                risk_eval['likelihood'],
                risk_eval['impact'],
                risk_eval['risk_level'],
                risk_eval['risk_score'],
                risk_eval['risk_category'],
                risk_eval['treatment_required']
            ]
            ws.append(row)
        
        self._adjust_column_widths(ws)
        self.worksheets['risk_evaluation'] = ws
    
    def _create_cybersecurity_goals_worksheet(self, analysis_data: Dict[str, Any]):
        """Create Cybersecurity Goals and Claims worksheet"""
        ws = self.workbook.create_sheet("Cybersecurity Goals")
        
        headers = [
            "Risk Threshold Level", "Risk Treatment Option", "ID", 
            "Cybersecurity Goal", "Cybersecurity Claim", "CAL", "Expected Functionality"
        ]
        ws.append(headers)
        self._style_headers(ws, len(headers))
        
        # Generate cybersecurity goals based on threats and mitigations
        threats = analysis_data.get('threats', [])
        mitigations = analysis_data.get('mitigations', [])
        
        for i, threat in enumerate(threats, 1):
            goal_data = self._generate_cybersecurity_goal(threat, mitigations)
            row = [
                goal_data['risk_threshold'],
                goal_data['treatment_option'],
                f"CG-{i:03d}",
                goal_data['goal'],
                goal_data['claim'],
                goal_data['cal'],
                goal_data['expected_functionality']
            ]
            ws.append(row)
        
        self._adjust_column_widths(ws)
        self.worksheets['cybersecurity_goals'] = ws
    
    def _create_summary_worksheet(self, analysis_data: Dict[str, Any], 
                                input_type: str, cross_ref_source: str):
        """Create executive summary worksheet"""
        ws = self.workbook.create_sheet("Executive Summary", 0)  # Make it first sheet
        
        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = "TARA Excel Report - Executive Summary"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Analysis details
        row = 3
        details = [
            ["Analysis Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Input Type", input_type.replace('_', ' ').title()],
            ["Cross-Reference Source", cross_ref_source.replace('_', ' ').title()],
            ["Total Threats Identified", len(analysis_data.get('threats', []))],
            ["Total Assets", len(analysis_data.get('assets', []))],
            ["Total Mitigations", len(analysis_data.get('mitigations', []))]
        ]
        
        for detail in details:
            ws[f'A{row}'] = detail[0]
            ws[f'B{row}'] = detail[1]
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Risk summary
        row += 2
        ws[f'A{row}'] = "Risk Summary"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 1
        
        risk_summary = self._generate_risk_summary(analysis_data)
        for risk_level, count in risk_summary.items():
            ws[f'A{row}'] = f"{risk_level} Risk Threats"
            ws[f'B{row}'] = count
            row += 1
        
        self._adjust_column_widths(ws)
        self.worksheets['summary'] = ws
    
    def _style_headers(self, worksheet, num_columns: int):
        """Apply consistent styling to worksheet headers"""
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col in range(1, num_columns + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    def _adjust_column_widths(self, worksheet):
        """Auto-adjust column widths based on content"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Helper methods for data analysis and generation
    
    def _generate_damage_scenario(self, threat: Dict, stakeholder: str) -> str:
        """Generate damage scenario description for a threat and stakeholder"""
        threat_name = threat.get('name', 'Unknown Threat')
        
        scenarios = {
            'End Users': f"Personal data exposure or service disruption due to {threat_name}",
            'Organization': f"Business impact and reputation damage from {threat_name}",
            'Third Parties': f"Supply chain or partner relationship impact from {threat_name}",
            'Regulators': f"Compliance violations and regulatory penalties due to {threat_name}"
        }
        
        return scenarios.get(stakeholder, f"Impact from {threat_name}")
    
    def _calculate_impact_analysis(self, threat: Dict) -> Dict[str, str]:
        """Calculate impact analysis for a threat"""
        severity = threat.get('severity', 'Medium').lower()
        
        impact_mapping = {
            'critical': {'safety': 'High', 'financial': 'High', 'operational': 'High', 'privacy': 'High'},
            'high': {'safety': 'Medium', 'financial': 'High', 'operational': 'Medium', 'privacy': 'High'},
            'medium': {'safety': 'Low', 'financial': 'Medium', 'operational': 'Medium', 'privacy': 'Medium'},
            'low': {'safety': 'Low', 'financial': 'Low', 'operational': 'Low', 'privacy': 'Low'}
        }
        
        impacts = impact_mapping.get(severity, impact_mapping['medium'])
        
        return {
            **impacts,
            'overall_level': severity.title(),
            'justification': f"Based on threat severity: {severity} and potential impact scope"
        }
    
    def _analyze_stride_threat(self, threat: Dict) -> Dict[str, str]:
        """Analyze threat against STRIDE categories"""
        threat_type = threat.get('type', '').lower()
        threat_desc = threat.get('description', '').lower()
        
        stride = {
            'spoofing': 'Not Applicable',
            'tampering': 'Not Applicable',
            'repudiation': 'Not Applicable',
            'information_disclosure': 'Not Applicable',
            'denial_of_service': 'Not Applicable',
            'elevation_of_privileges': 'Not Applicable'
        }
        
        # Simple keyword-based analysis
        if any(word in threat_desc for word in ['identity', 'authentication', 'imperson']):
            stride['spoofing'] = 'High Risk'
        
        if any(word in threat_desc for word in ['modify', 'alter', 'tamper', 'change']):
            stride['tampering'] = 'High Risk'
        
        if any(word in threat_desc for word in ['deny', 'log', 'audit', 'trace']):
            stride['repudiation'] = 'Medium Risk'
        
        if any(word in threat_desc for word in ['data', 'information', 'leak', 'exposure']):
            stride['information_disclosure'] = 'High Risk'
        
        if any(word in threat_desc for word in ['dos', 'denial', 'availability', 'service']):
            stride['denial_of_service'] = 'High Risk'
        
        if any(word in threat_desc for word in ['privilege', 'escalation', 'admin', 'root']):
            stride['elevation_of_privileges'] = 'High Risk'
        
        return stride
    
    def _generate_attack_path(self, threat: Dict, analysis_data: Dict) -> Dict[str, str]:
        """Generate attack path information"""
        return {
            'path_name': f"Attack Path for {threat.get('name', 'Unknown')}",
            'entry_point': 'External Network Interface',
            'target': threat.get('target', 'System Assets'),
            'steps': '1. Initial Access → 2. Persistence → 3. Privilege Escalation → 4. Impact'
        }
    
    def _assess_attack_feasibility(self, threat: Dict) -> Dict[str, str]:
        """Assess attack feasibility parameters"""
        severity = threat.get('severity', 'Medium').lower()
        
        feasibility_mapping = {
            'critical': {
                'elapsed_time': '< 1 day',
                'specialist_expertise': 'Layman',
                'knowledge_of_item': 'Public',
                'window_of_opportunity': 'Unlimited',
                'equipment': 'Standard',
                'attack_vector': 'Remote'
            },
            'high': {
                'elapsed_time': '< 1 week',
                'specialist_expertise': 'Proficient',
                'knowledge_of_item': 'Restricted',
                'window_of_opportunity': 'Moderate',
                'equipment': 'Specialized',
                'attack_vector': 'Adjacent Network'
            },
            'medium': {
                'elapsed_time': '< 1 month',
                'specialist_expertise': 'Expert',
                'knowledge_of_item': 'Confidential',
                'window_of_opportunity': 'Difficult',
                'equipment': 'Bespoke',
                'attack_vector': 'Local Network'
            },
            'low': {
                'elapsed_time': '> 6 months',
                'specialist_expertise': 'Multiple Expert',
                'knowledge_of_item': 'Strictly Confidential',
                'window_of_opportunity': 'Very Difficult',
                'equipment': 'Multiple Bespoke',
                'attack_vector': 'Physical'
            }
        }
        
        feasibility = feasibility_mapping.get(severity, feasibility_mapping['medium'])
        feasibility['summary_feasibility'] = f"{severity.title()} Feasibility"
        feasibility['risk_determination'] = f"{severity.title()} Risk"
        
        return feasibility
    
    def _evaluate_risk(self, threat: Dict) -> Dict[str, str]:
        """Evaluate risk parameters"""
        severity = threat.get('severity', 'Medium')
        
        risk_mapping = {
            'Critical': {'likelihood': 'Very High', 'impact': 'Critical', 'risk_score': '25'},
            'High': {'likelihood': 'High', 'impact': 'High', 'risk_score': '16'},
            'Medium': {'likelihood': 'Medium', 'impact': 'Medium', 'risk_score': '9'},
            'Low': {'likelihood': 'Low', 'impact': 'Low', 'risk_score': '4'}
        }
        
        risk_data = risk_mapping.get(severity, risk_mapping['Medium'])
        risk_data['risk_level'] = severity
        risk_data['risk_category'] = 'Cybersecurity Risk'
        risk_data['treatment_required'] = 'Yes' if severity in ['Critical', 'High'] else 'Optional'
        
        return risk_data
    
    def _generate_cybersecurity_goal(self, threat: Dict, mitigations: List[Dict]) -> Dict[str, str]:
        """Generate cybersecurity goal and claim"""
        severity = threat.get('severity', 'Medium')
        
        return {
            'risk_threshold': 'Medium' if severity in ['Critical', 'High'] else 'Low',
            'treatment_option': 'Mitigate',
            'goal': f"Prevent or mitigate {threat.get('name', 'security threat')}",
            'claim': f"System implements controls to address {threat.get('name', 'identified threat')}",
            'cal': 'CAL-2' if severity in ['Critical', 'High'] else 'CAL-1',
            'expected_functionality': 'Security controls operate as designed without impacting system functionality'
        }
    
    def _generate_risk_summary(self, analysis_data: Dict) -> Dict[str, int]:
        """Generate risk level summary"""
        threats = analysis_data.get('threats', [])
        summary = {'Critical': 0, 'High': 0, 'Medium': 0, 'Low': 0}
        
        for threat in threats:
            severity = threat.get('severity', 'Medium')
            if severity in summary:
                summary[severity] += 1
        
        return summary