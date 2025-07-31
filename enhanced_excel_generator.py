"""
Enhanced Excel Report Generator for TARA Analysis
"""

import pandas as pd
import os
import logging
from typing import Dict, List, Any, Optional
from datetime import datetime

class EnhancedTaraExcelGenerator:
    """Generate comprehensive TARA Excel reports with multiple worksheets"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
    
    def generate_excel_report(self, analysis_data: Dict[str, Any], 
                            input_type: str, 
                            cross_ref_source: str,
                            embed_assessment: Optional[Dict] = None) -> str:
        """Generate complete Excel report with all analysis data"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"TARA_Excel_Report_{timestamp}.xlsx"
            
            # Create uploads directory if it doesn't exist
            output_dir = 'uploads'
            os.makedirs(output_dir, exist_ok=True)
            
            filepath = os.path.join(output_dir, filename)
            
            # Create writer object
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Create single consolidated worksheet
                self._create_consolidated_sheet(writer, analysis_data, input_type, cross_ref_source, embed_assessment)
            
            self.logger.info(f"Excel report generated: {filepath}")
            return filepath
            
        except Exception as e:
            self.logger.error(f"Failed to generate Excel report: {e}")
            raise
    
    def _create_consolidated_sheet(self, writer, analysis_data: Dict, input_type: str, cross_ref_source: str, embed_assessment: Optional[Dict]):
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        # Create all required worksheets as specified
        self._create_required_worksheets(writer.book)
        
        # Create the main worksheet
        ws = writer.book.create_sheet('TARA Analysis Report')
        writer.book.active = ws
        
        # Define color schemes for each section
        colors = {
            'assets': 'E7E6E6',
            'damage_scenario': 'E2EFDA', 
            'impact_analysis': 'FFF2CC',
            'cybersecurity_controls': 'F2F2F2',
            'threat_scenarios': 'DEEAF6',
            'attack_path': 'FCE4D6',
            'attack_feasibility': 'EDEDED',
            'cybersecurity_goals': 'E1D5E7'
        }
        
        # Define section headers and their columns
        sections = {
            'Assets': ['Asset ID', 'Asset Name', 'Security Property Loss'],
            'Damage Scenario': ['Stakeholder', 'Damage Scenario Description'],
            'Impact Analysis': ['Impact Category', 'Impact Level', 'Impact Description'],
            'Cybersecurity Controls': ['Control ID', 'Control Name', 'Control Description'],
            'Threat Scenarios': ['Threat ID', 'Threat Name', 'STRIDE Category'],
            'Attack Path': ['Attack Vector', 'Attack Steps', 'Prerequisites'],
            'Attack Feasibility Assessment': ['Feasibility Factor', 'Rating', 'Justification'],
            'Cybersecurity Goals and Claims': ['Goal ID', 'Security Goal', 'Claim Statement']
        }
        
        # Start position
        current_col = 1
        
        # Create each section
        for section_name, columns in sections.items():
            color_key = section_name.lower().replace(' ', '_').replace('_and_', '_')
            if color_key not in colors:
                color_key = 'assets'  # Default color
            
            fill_color = colors[color_key]
            
            # Create section header
            self._create_section_header(ws, section_name, current_col, len(columns), fill_color)
            
            # Create column headers
            for i, col_name in enumerate(columns):
                cell = ws.cell(row=2, column=current_col + i)
                cell.value = col_name
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                cell.font = Font(bold=True, size=10)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            # Fill section data
            self._fill_section_data(ws, section_name, analysis_data, current_col, fill_color, embed_assessment)
            
            # Move to next section
            current_col += len(columns) + 1  # Add gap between sections
        
        # Auto-fit column widths
        from openpyxl.utils import get_column_letter
        from openpyxl.cell.cell import MergedCell
        
        for col_idx in range(1, ws.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                try:
                    # Skip merged cells and check for actual content
                    if not isinstance(cell, MergedCell) and cell.value is not None:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            adjusted_width = min(max(max_length + 2, 12), 50)  # Min 12, max 50
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_section_header(self, ws, section_name, start_col, num_cols, fill_color):
        """Create merged header for section"""
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        # Merge cells for header
        if num_cols > 1:
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=start_col + num_cols - 1)
        
        # Style the header
        header_cell = ws.cell(row=1, column=start_col)
        header_cell.value = section_name
        header_cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        header_cell.font = Font(bold=True, size=12)
        header_cell.alignment = Alignment(horizontal='center', vertical='center')
        header_cell.border = Border(
            left=Side(style='thick'),
            right=Side(style='thick'),
            top=Side(style='thick'),
            bottom=Side(style='thick')
        )
    
    def _fill_section_data(self, ws, section_name, analysis_data, start_col, fill_color, embed_assessment):
        """Fill data for each section"""
        from openpyxl.styles import PatternFill, Alignment, Border, Side
        
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        row = 3  # Start data from row 3
        
        if section_name == 'Assets':
            assets = analysis_data.get('assets', [])
            for i, asset in enumerate(assets):
                ws.cell(row=row + i, column=start_col).value = f"A{i+1:03d}"
                ws.cell(row=row + i, column=start_col + 1).value = asset.get('name', 'Unknown')
                ws.cell(row=row + i, column=start_col + 2).value = asset.get('criticality', 'Medium')
                
                for col in range(3):
                    cell = ws.cell(row=row + i, column=start_col + col)
                    cell.fill = fill
                    cell.border = border
                    cell.alignment = Alignment(vertical='center')
        
        elif section_name == 'Damage Scenario':
            threats = analysis_data.get('threats', [])
            stakeholders = ['Users', 'Organization', 'Partners', 'Regulators']
            
            scenario_row = row
            for threat in threats[:10]:  # Limit to first 10 threats
                for stakeholder in stakeholders:
                    if scenario_row - row < 20:  # Limit total scenarios
                        ws.cell(row=scenario_row, column=start_col).value = stakeholder
                        scenario_desc = self._generate_damage_scenario(threat, stakeholder)
                        ws.cell(row=scenario_row, column=start_col + 1).value = scenario_desc
                        
                        for col in range(2):
                            cell = ws.cell(row=scenario_row, column=start_col + col)
                            cell.fill = fill
                            cell.border = border
                            cell.alignment = Alignment(vertical='center', wrap_text=True)
                        
                        scenario_row += 1
        
        elif section_name == 'Impact Analysis':
            threats = analysis_data.get('threats', [])
            for i, threat in enumerate(threats[:15]):  # Limit to 15 threats
                impact = self._calculate_impact_analysis(threat)
                ws.cell(row=row + i, column=start_col).value = impact.get('category', 'Operational')
                ws.cell(row=row + i, column=start_col + 1).value = impact.get('level', 'Medium')
                ws.cell(row=row + i, column=start_col + 2).value = impact.get('description', 'Standard impact assessment')
                
                for col in range(3):
                    cell = ws.cell(row=row + i, column=start_col + col)
                    cell.fill = fill
                    cell.border = border
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
        
        elif section_name == 'Cybersecurity Controls':
            controls = analysis_data.get('mitigations', [])
            if embed_assessment:
                controls.extend(embed_assessment.get('controls', []))
            
            for i, control in enumerate(controls[:20]):  # Limit to 20 controls
                ws.cell(row=row + i, column=start_col).value = f"C{i+1:03d}"
                ws.cell(row=row + i, column=start_col + 1).value = control.get('name', control.get('title', 'Security Control'))
                ws.cell(row=row + i, column=start_col + 2).value = control.get('description', 'Standard security control')
                
                for col in range(3):
                    cell = ws.cell(row=row + i, column=start_col + col)
                    cell.fill = fill
                    cell.border = border
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
        
        elif section_name == 'Threat Scenarios':
            threats = analysis_data.get('threats', [])
            for i, threat in enumerate(threats[:15]):
                stride = self._analyze_stride_threat(threat)
                ws.cell(row=row + i, column=start_col).value = f"T{i+1:03d}"
                ws.cell(row=row + i, column=start_col + 1).value = threat.get('name', threat.get('title', 'Security Threat'))
                ws.cell(row=row + i, column=start_col + 2).value = stride.get('primary_category', 'Information Disclosure')
                
                for col in range(3):
                    cell = ws.cell(row=row + i, column=start_col + col)
                    cell.fill = fill
                    cell.border = border
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
        
        elif section_name == 'Attack Path':
            threats = analysis_data.get('threats', [])
            for i, threat in enumerate(threats[:12]):
                attack_path = self._generate_attack_path(threat)
                ws.cell(row=row + i, column=start_col).value = attack_path.get('vector', 'Network')
                ws.cell(row=row + i, column=start_col + 1).value = attack_path.get('steps', 'Multi-step attack sequence')
                ws.cell(row=row + i, column=start_col + 2).value = attack_path.get('prerequisites', 'Network access required')
                
                for col in range(3):
                    cell = ws.cell(row=row + i, column=start_col + col)
                    cell.fill = fill
                    cell.border = border
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
        
        elif section_name == 'Attack Feasibility Assessment':
            threats = analysis_data.get('threats', [])
            feasibility_factors = ['Technical Skill', 'Resources Required', 'Detection Likelihood']
            
            factor_row = row
            for threat in threats[:8]:  # Limit threats
                for factor in feasibility_factors:
                    if factor_row - row < 24:  # Limit total assessments
                        feasibility = self._assess_attack_feasibility(threat)
                        ws.cell(row=factor_row, column=start_col).value = factor
                        ws.cell(row=factor_row, column=start_col + 1).value = feasibility.get('rating', 'Medium')
                        ws.cell(row=factor_row, column=start_col + 2).value = feasibility.get('justification', 'Standard assessment')
                        
                        for col in range(3):
                            cell = ws.cell(row=factor_row, column=start_col + col)
                            cell.fill = fill
                            cell.border = border
                            cell.alignment = Alignment(vertical='center', wrap_text=True)
                        
                        factor_row += 1
        
        elif section_name == 'Cybersecurity Goals and Claims':
            threats = analysis_data.get('threats', [])
            mitigations = analysis_data.get('mitigations', [])
            
            for i, threat in enumerate(threats[:12]):
                goal = self._generate_cybersecurity_goal(threat, mitigations)
                ws.cell(row=row + i, column=start_col).value = f"G{i+1:03d}"
                ws.cell(row=row + i, column=start_col + 1).value = goal.get('goal', 'Ensure system security')
                ws.cell(row=row + i, column=start_col + 2).value = goal.get('claim', 'Security controls effectively mitigate threats')
                
                for col in range(3):
                    cell = ws.cell(row=row + i, column=start_col + col)
                    cell.fill = fill
                    cell.border = border
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
    
    def _create_executive_summary(self, writer, analysis_data: Dict, input_type: str, cross_ref_source: str):
        """Create executive summary worksheet"""
        summary_data = {
            'Analysis Details': [
                'Analysis Date',
                'Input Type',
                'Cross-Reference Source',
                'Total Threats',
                'Total Assets',
                'Total Mitigations'
            ],
            'Values': [
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                input_type.replace('_', ' ').title(),
                cross_ref_source.replace('_', ' ').title(),
                len(analysis_data.get('threats', [])),
                len(analysis_data.get('assets', [])),
                len(analysis_data.get('mitigations', []))
            ]
        }
        
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name='Executive Summary', index=False)
    
    def _create_assets_sheet(self, writer, analysis_data: Dict):
        """Create Assets worksheet"""
        assets = analysis_data.get('assets', [])
        
        assets_data = []
        for i, asset in enumerate(assets, 1):
            assets_data.append({
                'Asset ID': f"A{i:03d}",
                'Asset': asset.get('name', 'Unknown Asset'),
                'Security Property Loss': 'Confidentiality, Integrity, Availability'
            })
        
        df_assets = pd.DataFrame(assets_data)
        df_assets.to_excel(writer, sheet_name='Assets', index=False)
    
    def _create_damage_scenarios_sheet(self, writer, analysis_data: Dict):
        """Create Damage Scenarios worksheet"""
        threats = analysis_data.get('threats', [])
        stakeholders = ['End Users', 'Organization', 'Third Parties', 'Regulators']
        
        scenarios_data = []
        for threat in threats:
            for stakeholder in stakeholders:
                scenario = self._generate_damage_scenario(threat, stakeholder)
                scenarios_data.append({
                    'Stakeholder': stakeholder,
                    'Damage Scenario Description': scenario
                })
        
        df_scenarios = pd.DataFrame(scenarios_data)
        df_scenarios.to_excel(writer, sheet_name='Damage Scenarios', index=False)
    
    def _create_impact_analysis_sheet(self, writer, analysis_data: Dict):
        """Create Impact Analysis worksheet"""
        threats = analysis_data.get('threats', [])
        
        impact_data = []
        for threat in threats:
            impact = self._calculate_impact_analysis(threat)
            impact_data.append({
                'Threat ID': threat.get('id', 'T-XXX'),
                'Safety': impact['safety'],
                'Financial': impact['financial'],
                'Operational': impact['operational'],
                'Privacy': impact['privacy'],
                'Impact Level': impact['overall_level'],
                'Justification': impact['justification']
            })
        
        df_impact = pd.DataFrame(impact_data)
        df_impact.to_excel(writer, sheet_name='Impact Analysis', index=False)
    
    def _create_cybersecurity_controls_sheet(self, writer, analysis_data: Dict, embed_assessment: Optional[Dict]):
        """Create Cybersecurity Controls worksheet"""
        controls_data = []
        
        # Add existing mitigations
        mitigations = analysis_data.get('mitigations', [])
        for mitigation in mitigations:
            controls_data.append({
                'ID': mitigation.get('id', 'C-XXX'),
                'Cybersecurity Control': mitigation.get('description', mitigation.get('name', 'Unknown Control'))
            })
        
        # Add EMBED controls if available
        if embed_assessment and 'recommended_controls' in embed_assessment:
            for control in embed_assessment['recommended_controls']:
                controls_data.append({
                    'ID': control.get('id', 'EMBED-XXX'),
                    'Cybersecurity Control': control.get('control', control.get('description', 'EMBED Control'))
                })
        
        df_controls = pd.DataFrame(controls_data)
        df_controls.to_excel(writer, sheet_name='Cybersecurity Controls', index=False)
    
    def _create_threat_scenarios_sheet(self, writer, analysis_data: Dict):
        """Create Threat Scenarios worksheet (STRIDE)"""
        threats = analysis_data.get('threats', [])
        
        stride_data = []
        for threat in threats:
            stride = self._analyze_stride_threat(threat)
            stride_data.append({
                'Threat ID': threat.get('id', 'T-XXX'),
                'Spoofing': stride['spoofing'],
                'Tampering': stride['tampering'],
                'Repudiation': stride['repudiation'],
                'Information Disclosure': stride['information_disclosure'],
                'Denial of Service': stride['denial_of_service'],
                'Elevation of Privileges': stride['elevation_of_privileges']
            })
        
        df_stride = pd.DataFrame(stride_data)
        df_stride.to_excel(writer, sheet_name='Threat Scenarios', index=False)
    
    def _create_attack_paths_sheet(self, writer, analysis_data: Dict):
        """Create Attack Paths worksheet"""
        threats = analysis_data.get('threats', [])
        
        attack_paths_data = []
        for threat in threats:
            attack_path = self._generate_attack_path(threat)
            attack_paths_data.append({
                'Threat ID': threat.get('id', 'T-XXX'),
                'Attack Path': attack_path['path_name'],
                'Entry Point': attack_path['entry_point'],
                'Target Asset': attack_path['target'],
                'Attack Steps': attack_path['steps']
            })
        
        df_paths = pd.DataFrame(attack_paths_data)
        df_paths.to_excel(writer, sheet_name='Attack Paths', index=False)
    
    def _create_attack_feasibility_sheet(self, writer, analysis_data: Dict):
        """Create Attack Feasibility Assessment worksheet"""
        threats = analysis_data.get('threats', [])
        
        feasibility_data = []
        for threat in threats:
            feasibility = self._assess_attack_feasibility(threat)
            feasibility_data.append({
                'Threat ID': threat.get('id', 'T-XXX'),
                'Elapsed Time': feasibility['elapsed_time'],
                'Specialist Expertise': feasibility['specialist_expertise'],
                'Knowledge of Item': feasibility['knowledge_of_item'],
                'Window of Opportunity': feasibility['window_of_opportunity'],
                'Equipment': feasibility['equipment'],
                'Attack Vector': feasibility['attack_vector'],
                'Summary Attack Feasibility': feasibility['summary_feasibility'],
                'Risk Determination': feasibility['risk_determination']
            })
        
        df_feasibility = pd.DataFrame(feasibility_data)
        df_feasibility.to_excel(writer, sheet_name='Attack Feasibility', index=False)
    
    def _create_risk_evaluation_sheet(self, writer, analysis_data: Dict):
        """Create Risk Evaluation worksheet"""
        threats = analysis_data.get('threats', [])
        
        risk_data = []
        for threat in threats:
            risk = self._evaluate_risk(threat)
            risk_data.append({
                'Threat ID': threat.get('id', 'T-XXX'),
                'Likelihood': risk['likelihood'],
                'Impact': risk['impact'],
                'Risk Level': risk['risk_level'],
                'Risk Score': risk['risk_score'],
                'Risk Category': risk['risk_category'],
                'Treatment Required': risk['treatment_required']
            })
        
        df_risk = pd.DataFrame(risk_data)
        df_risk.to_excel(writer, sheet_name='Risk Evaluation', index=False)
    
    def _create_cybersecurity_goals_sheet(self, writer, analysis_data: Dict):
        """Create Cybersecurity Goals and Claims worksheet"""
        threats = analysis_data.get('threats', [])
        mitigations = analysis_data.get('mitigations', [])
        
        goals_data = []
        for i, threat in enumerate(threats, 1):
            goal = self._generate_cybersecurity_goal(threat, mitigations)
            goals_data.append({
                'Risk Threshold Level': goal['risk_threshold'],
                'Risk Treatment Option': goal['treatment_option'],
                'ID': f"CG-{i:03d}",
                'Cybersecurity Goal': goal['goal'],
                'Cybersecurity Claim': goal['claim'],
                'CAL': goal['cal'],
                'Expected Functionality': goal['expected_functionality']
            })
        
        df_goals = pd.DataFrame(goals_data)
        df_goals.to_excel(writer, sheet_name='Cybersecurity Goals', index=False)
    
    # Helper methods for data analysis
    
    def _generate_damage_scenario(self, threat: Dict, stakeholder: str) -> str:
        """Generate damage scenario description"""
        threat_name = threat.get('name', 'Unknown Threat')
        
        scenarios = {
            'End Users': f"Personal data exposure or service disruption due to {threat_name}",
            'Organization': f"Business impact and reputation damage from {threat_name}",
            'Third Parties': f"Supply chain or partner relationship impact from {threat_name}",
            'Regulators': f"Compliance violations and regulatory penalties due to {threat_name}"
        }
        
        return scenarios.get(stakeholder, f"Impact from {threat_name}")
    
    def _calculate_impact_analysis(self, threat: Dict) -> Dict[str, str]:
        """Calculate impact analysis for a threat"""
        severity = threat.get('severity', 'Medium').lower()
        
        impact_mapping = {
            'critical': {'safety': 'High', 'financial': 'High', 'operational': 'High', 'privacy': 'High'},
            'high': {'safety': 'Medium', 'financial': 'High', 'operational': 'Medium', 'privacy': 'High'},
            'medium': {'safety': 'Low', 'financial': 'Medium', 'operational': 'Medium', 'privacy': 'Medium'},
            'low': {'safety': 'Low', 'financial': 'Low', 'operational': 'Low', 'privacy': 'Low'}
        }
        
        impacts = impact_mapping.get(severity, impact_mapping['medium'])
        
        return {
            **impacts,
            'overall_level': severity.title(),
            'justification': f"Based on threat severity: {severity} and potential impact scope"
        }
    
    def _analyze_stride_threat(self, threat: Dict) -> Dict[str, str]:
        """Analyze threat against STRIDE categories"""
        threat_desc = threat.get('description', '').lower()
        
        stride = {
            'spoofing': 'Not Applicable',
            'tampering': 'Not Applicable',
            'repudiation': 'Not Applicable',
            'information_disclosure': 'Not Applicable',
            'denial_of_service': 'Not Applicable',
            'elevation_of_privileges': 'Not Applicable'
        }
        
        if any(word in threat_desc for word in ['identity', 'authentication', 'imperson']):
            stride['spoofing'] = 'High Risk'
        
        if any(word in threat_desc for word in ['modify', 'alter', 'tamper', 'change']):
            stride['tampering'] = 'High Risk'
        
        if any(word in threat_desc for word in ['deny', 'log', 'audit', 'trace']):
            stride['repudiation'] = 'Medium Risk'
        
        if any(word in threat_desc for word in ['data', 'information', 'leak', 'exposure']):
            stride['information_disclosure'] = 'High Risk'
        
        if any(word in threat_desc for word in ['dos', 'denial', 'availability', 'service']):
            stride['denial_of_service'] = 'High Risk'
        
        if any(word in threat_desc for word in ['privilege', 'escalation', 'admin', 'root']):
            stride['elevation_of_privileges'] = 'High Risk'
        
        return stride
    
    def _generate_attack_path(self, threat: Dict) -> Dict[str, str]:
        """Generate attack path information"""
        return {
            'path_name': f"Attack Path for {threat.get('name', 'Unknown')}",
            'entry_point': 'External Network Interface',
            'target': threat.get('target', 'System Assets'),
            'steps': '1. Initial Access → 2. Persistence → 3. Privilege Escalation → 4. Impact'
        }
    
    def _assess_attack_feasibility(self, threat: Dict) -> Dict[str, str]:
        """Assess attack feasibility parameters"""
        severity = threat.get('severity', 'Medium').lower()
        
        feasibility_mapping = {
            'critical': {
                'elapsed_time': '< 1 day',
                'specialist_expertise': 'Layman',
                'knowledge_of_item': 'Public',
                'window_of_opportunity': 'Unlimited',
                'equipment': 'Standard',
                'attack_vector': 'Remote'
            },
            'high': {
                'elapsed_time': '< 1 week',
                'specialist_expertise': 'Proficient',
                'knowledge_of_item': 'Restricted',
                'window_of_opportunity': 'Moderate',
                'equipment': 'Specialized',
                'attack_vector': 'Adjacent Network'
            },
            'medium': {
                'elapsed_time': '< 1 month',
                'specialist_expertise': 'Expert',
                'knowledge_of_item': 'Confidential',
                'window_of_opportunity': 'Difficult',
                'equipment': 'Bespoke',
                'attack_vector': 'Local Network'
            },
            'low': {
                'elapsed_time': '> 6 months',
                'specialist_expertise': 'Multiple Expert',
                'knowledge_of_item': 'Strictly Confidential',
                'window_of_opportunity': 'Very Difficult',
                'equipment': 'Multiple Bespoke',
                'attack_vector': 'Physical'
            }
        }
        
        feasibility = feasibility_mapping.get(severity, feasibility_mapping['medium'])
        feasibility['summary_feasibility'] = f"{severity.title()} Feasibility"
        feasibility['risk_determination'] = f"{severity.title()} Risk"
        
        return feasibility
    
    def _evaluate_risk(self, threat: Dict) -> Dict[str, str]:
        """Evaluate risk parameters"""
        severity = threat.get('severity', 'Medium')
        
        risk_mapping = {
            'Critical': {'likelihood': 'Very High', 'impact': 'Critical', 'risk_score': '25'},
            'High': {'likelihood': 'High', 'impact': 'High', 'risk_score': '16'},
            'Medium': {'likelihood': 'Medium', 'impact': 'Medium', 'risk_score': '9'},
            'Low': {'likelihood': 'Low', 'impact': 'Low', 'risk_score': '4'}
        }
        
        risk_data = risk_mapping.get(severity, risk_mapping['Medium'])
        risk_data['risk_level'] = severity
        risk_data['risk_category'] = 'Cybersecurity Risk'
        risk_data['treatment_required'] = 'Yes' if severity in ['Critical', 'High'] else 'Optional'
        
        return risk_data
    
    def _generate_cybersecurity_goal(self, threat: Dict, mitigations: List[Dict]) -> Dict[str, str]:
        """Generate cybersecurity goal and claim"""
        severity = threat.get('severity', 'Medium')
        
        return {
            'risk_threshold': 'Medium' if severity in ['Critical', 'High'] else 'Low',
            'treatment_option': 'Mitigate',
            'goal': f"Prevent or mitigate {threat.get('name', 'security threat')}",
            'claim': f"System implements controls to address {threat.get('name', 'identified threat')}",
            'cal': 'CAL-2' if severity in ['Critical', 'High'] else 'CAL-1',
            'expected_functionality': 'Security controls operate as designed without impacting system functionality'
        }
    
    def _create_required_worksheets(self, wb):
        """Create all required worksheets with specified column headers"""
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        # General Information sheet with specific formatting
        general_info = wb.create_sheet("General information")
        self._format_general_info_sheet(general_info)
        
        # Item Definition sheet with specific formatting
        item_def = wb.create_sheet("Item definition")
        self._format_item_definition_sheet(item_def)
        
        # Threat Model sheet with specific formatting
        threat_model = wb.create_sheet("Item definition - Threat model")
        self._format_threat_model_sheet(threat_model)
        
        # Asset Identification sheet with specific formatting
        asset_id_sheet = wb.create_sheet("Asset identification")
        self._format_asset_identification_sheet(asset_id_sheet)
        
        # TARA sheet (main analysis sheet) with specific formatting
        tara_sheet = wb.create_sheet("TARA")
        self._format_tara_sheet(tara_sheet)
        
        # Matrices and Guidelines sheet with specific formatting
        matrix_sheet = wb.create_sheet("Matrices and guidlines")
        self._format_matrices_guidelines_sheet(matrix_sheet)
        
        # Template Revision History sheet
        rev_sheet = wb.create_sheet("Template revision history")
        rev_sheet.append([
            "Version", "Date", "Author", "Changes"
        ])
    
    def _format_general_info_sheet(self, ws):
        """Format the General Information sheet with exact specifications"""
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        # Define styles
        header_fill = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")
        bold_border = Border(
            left=Side(style='thick', color='000000'),
            right=Side(style='thick', color='000000'),
            top=Side(style='thick', color='000000'),
            bottom=Side(style='thick', color='000000')
        )
        
        # Set row heights
        ws.row_dimensions[1].height = 49.5
        ws.row_dimensions[2].height = 29.3
        ws.row_dimensions[3].height = 29.3
        for row in range(6, 20):
            ws.row_dimensions[row].height = 22.5
        
        # Title font styles
        title_font = Font(name='Calibri', size=20, bold=True, color='000000')
        project_font = Font(name='Calibri', size=14, bold=True, color='000000')
        responsible_font = Font(name='Calibri', size=14, bold=True, color='000000')
        revision_font = Font(name='Calibri', size=11, bold=True, color='000000')
        header_font = Font(name='Calibri', size=11, bold=True, color='000000')
        
        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')
        
        # Lines 1-3: Merge cells B-H and add content with styling
        ws.merge_cells('B1:H1')
        ws['B1'] = "Threat analysis and risk assessment (TARA)"
        ws['B1'].font = title_font
        ws['B1'].alignment = center_alignment
        ws['B1'].fill = header_fill
        
        ws.merge_cells('B2:H2')
        ws['B2'] = "Project:"
        ws['B2'].font = project_font
        ws['B2'].alignment = center_alignment
        ws['B2'].fill = header_fill
        
        ws.merge_cells('B3:H3')
        ws['B3'] = "Responsible:"
        ws['B3'].font = responsible_font
        ws['B3'].alignment = center_alignment
        ws['B3'].fill = header_fill
        
        # Apply bold borders to header box B1:H3
        for row in range(1, 4):
            for col in range(2, 9):  # B to H
                cell = ws.cell(row=row, column=col)
                cell.border = bold_border
        
        # Line 5: "Revision history" in column B
        ws['B5'] = "Revision history"
        ws['B5'].font = revision_font
        ws['B5'].alignment = left_alignment
        
        # Lines 6-19: Revision history table
        # Set column widths as specified
        ws.column_dimensions['B'].width = 18.26  # Revision number
        ws.column_dimensions['C'].width = 16.11  # Revision reason
        ws.column_dimensions['D'].width = 26.26  # Author(s)
        ws.column_dimensions['E'].width = 51.11  # Description
        ws.column_dimensions['F'].width = 15.11  # Date
        ws.column_dimensions['G'].width = 26.26  # Approver(s)
        ws.column_dimensions['H'].width = 18.26  # Approval date
        
        # Row 6: Headers with formatting
        headers = [
            "Revision number", "Revision reason", "Author(s)", 
            "Description", "Date", "Approver(s)", "Approval date"
        ]
        
        for col_idx, header in enumerate(headers, start=2):  # Start from column B
            cell = ws.cell(row=6, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.alignment = left_alignment
            cell.fill = header_fill
            cell.border = bold_border
        
        # Rows 7-19: Empty data rows with bold borders
        for row in range(7, 20):
            for col in range(2, 9):  # B to H
                cell = ws.cell(row=row, column=col)
                cell.border = bold_border
    
    def _format_item_definition_sheet(self, ws):
        """Format the Item Definition sheet with exact specifications"""
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        # Define styles
        header_fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
        section_fill = PatternFill(start_color="D0CECE", end_color="D0CECE", fill_type="solid")
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Font styles
        title_font = Font(name='Calibri', size=20, bold=True, color='000000')
        header_font = Font(name='Calibri', size=11, color='000000')
        
        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')
        
        # Set column widths as specified
        ws.column_dimensions['A'].width = 8.58
        ws.column_dimensions['B'].width = 8.84
        ws.column_dimensions['C'].width = 101.58
        ws.column_dimensions['D'].width = 8.58
        ws.column_dimensions['E'].width = 150.58
        ws.column_dimensions['F'].width = 8.58
        
        # Rows 1-2: Merge B-E for "Item Definition" title
        ws.merge_cells('B1:E2')
        ws['B1'] = "Item Definition"
        ws['B1'].font = title_font
        ws['B1'].alignment = center_alignment
        ws['B1'].fill = header_fill
        
        # Apply borders to merged cells B1:E2
        for row in range(1, 3):
            for col in range(2, 6):  # B to E
                cell = ws.cell(row=row, column=col)
                cell.border = border
        
        # Row 4-25: ID and Item functions section
        # Headers in row 4
        ws['B4'] = "ID"
        ws['B4'].font = header_font
        ws['B4'].alignment = left_alignment
        ws['B4'].fill = section_fill
        ws['B4'].border = border
        
        ws['C4'] = "Item functions/Technical approach"
        ws['C4'].font = header_font
        ws['C4'].alignment = left_alignment
        ws['C4'].fill = section_fill
        ws['C4'].border = border
        
        # Preliminary architecture header in E4
        ws['E4'] = "Preliminary architecture"
        ws['E4'].font = header_font
        ws['E4'].alignment = center_alignment
        ws['E4'].fill = section_fill
        ws['E4'].border = border
        
        # Apply borders to data rows 5-25 for columns B, C, E
        for row in range(5, 26):
            for col in [2, 3, 5]:  # B, C, E
                cell = ws.cell(row=row, column=col)
                cell.border = border
        
        # Row 27-34: Operational environment section
        ws['B27'] = "ID"
        ws['B27'].font = header_font
        ws['B27'].alignment = left_alignment
        ws['B27'].fill = section_fill
        ws['B27'].border = border
        
        ws['C27'] = "Operational environment"
        ws['C27'].font = header_font
        ws['C27'].alignment = left_alignment
        ws['C27'].fill = section_fill
        ws['C27'].border = border
        
        # Module overview header in E27
        ws['E27'] = "Module overview"
        ws['E27'].font = header_font
        ws['E27'].alignment = center_alignment
        ws['E27'].fill = section_fill
        ws['E27'].border = border
        
        # Apply borders to rows 28-34 for columns B, C
        for row in range(28, 35):
            for col in [2, 3]:  # B, C
                cell = ws.cell(row=row, column=col)
                cell.border = border
        
        # Row 36-43: Assumptions section
        ws['B36'] = "ID"
        ws['B36'].font = header_font
        ws['B36'].alignment = left_alignment
        ws['B36'].fill = section_fill
        ws['B36'].border = border
        
        ws['C36'] = "Assumptions"
        ws['C36'].font = header_font
        ws['C36'].alignment = left_alignment
        ws['C36'].fill = section_fill
        ws['C36'].border = border
        
        # Apply borders to rows 37-43 for columns B, C
        for row in range(37, 44):
            for col in [2, 3]:  # B, C
                cell = ws.cell(row=row, column=col)
                cell.border = border
        
        # Apply borders to column E for rows 28-43 (Module overview section)
        for row in range(28, 44):
            cell = ws.cell(row=row, column=5)  # Column E
            cell.border = border
        
        # Row 45-65: Item Boundary section
        ws.merge_cells('B45:E45')
        ws['B45'] = "Item Boundary"
        ws['B45'].font = header_font
        ws['B45'].alignment = center_alignment
        ws['B45'].fill = section_fill
        ws['B45'].border = border
        
        # Apply borders to merged header B45:E45
        for col in range(2, 6):  # B to E
            cell = ws.cell(row=45, column=col)
            cell.border = border
        
        # Apply outer borders only to rows 46-65 for columns B-E (no internal lines)
        # Top border for row 46
        for col in range(2, 6):  # B to E
            cell = ws.cell(row=46, column=col)
            cell.border = Border(top=Side(style='thin', color='000000'))
        
        # Bottom border for row 65
        for col in range(2, 6):  # B to E
            cell = ws.cell(row=65, column=col)
            cell.border = Border(bottom=Side(style='thin', color='000000'))
        
        # Left border for column B (rows 46-65)
        for row in range(46, 66):
            cell = ws.cell(row=row, column=2)  # Column B
            current_border = cell.border
            cell.border = Border(
                left=Side(style='thin', color='000000'),
                top=current_border.top,
                bottom=current_border.bottom
            )
        
        # Right border for column E (rows 46-65)
        for row in range(46, 66):
            cell = ws.cell(row=row, column=5)  # Column E
            current_border = cell.border
            cell.border = Border(
                right=Side(style='thin', color='000000'),
                top=current_border.top,
                bottom=current_border.bottom
            )
    
    def _format_threat_model_sheet(self, ws):
        """Format the Item definition - Threat model sheet"""
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        # Define styles
        header_fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
        section_fill = PatternFill(start_color="D0CECE", end_color="D0CECE", fill_type="solid")
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        title_font = Font(name='Calibri', size=16, color='000000')
        header_font = Font(name='Calibri', size=11, color='000000')
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Lines 1-2: Merge B-D for title
        ws.merge_cells('B1:D2')
        ws['B1'] = "Item definition – Threat model"
        ws['B1'].font = title_font
        ws['B1'].alignment = center_alignment
        ws['B1'].fill = header_fill
        
        # Apply borders to merged cells B1:D2
        for row in range(1, 3):
            for col in range(2, 5):  # B to D
                cell = ws.cell(row=row, column=col)
                cell.border = border
        
        # Lines 4-32: Column headers
        ws['B4'] = "Threat model – ECU 01"
        ws['B4'].font = header_font
        ws['B4'].alignment = center_alignment
        ws['B4'].fill = section_fill
        ws['B4'].border = border
        
        ws['D4'] = "Threat model – ECU 02"
        ws['D4'].font = header_font
        ws['D4'].alignment = center_alignment
        ws['D4'].fill = section_fill
        ws['D4'].border = border
        
        # Apply borders to data rows 5-32 for columns B and D
        for row in range(5, 33):
            for col in [2, 4]:  # B, D
                cell = ws.cell(row=row, column=col)
                cell.border = border
    
    def _format_asset_identification_sheet(self, ws):
        """Format the Asset identification sheet"""
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        # Define styles and colors
        main_header_fill = PatternFill(start_color="9BC2E6", end_color="9BC2E6", fill_type="solid")
        derivation_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        security_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        cyber_assets_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
        approach_fill = PatternFill(start_color="AEAAAA", end_color="AEAAAA", fill_type="solid")
        data_fill = PatternFill(start_color="D0CECE", end_color="D0CECE", fill_type="solid")
        
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        title_font = Font(name='Calibri', size=16, color='000000')
        header_font = Font(name='Calibri', size=11, color='000000')
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Set column widths
        ws.column_dimensions['A'].width = 8.58
        ws.column_dimensions['B'].width = 19.68
        ws.column_dimensions['C'].width = 42
        ws.column_dimensions['D'].width = 22
        ws.column_dimensions['E'].width = 22
        ws.column_dimensions['F'].width = 22
        ws.column_dimensions['G'].width = 28.68
        ws.column_dimensions['H'].width = 18.58
        
        # Rows 1-2: Main header B-H merged
        ws.merge_cells('B1:H2')
        ws['B1'] = "Asset identification process"
        ws['B1'].font = title_font
        ws['B1'].alignment = center_alignment
        ws['B1'].fill = main_header_fill
        
        # Apply borders to main header
        for row in range(1, 3):
            for col in range(2, 9):  # B to H
                cell = ws.cell(row=row, column=col)
                cell.border = border
        
        # Row 3: Section headers
        # B-C merged: "Derivation of a candidate asset"
        ws.merge_cells('B3:C3')
        ws['B3'] = "Derivation of a candidate asset"
        ws['B3'].font = header_font
        ws['B3'].alignment = center_alignment
        ws['B3'].fill = derivation_fill
        
        # D-F merged: "Determination of security property loss"
        ws.merge_cells('D3:F3')
        ws['D3'] = "Determination of security property loss"
        ws['D3'].font = header_font
        ws['D3'].alignment = center_alignment
        ws['D3'].fill = security_fill
        
        # G-H merged: "List of cybersecurity assets"
        ws.merge_cells('G3:H3')
        ws['G3'] = "List of cybersecurity assets"
        ws['G3'].font = header_font
        ws['G3'].alignment = center_alignment
        ws['G3'].fill = cyber_assets_fill
        
        # Apply borders to row 3
        for col in range(2, 9):
            cell = ws.cell(row=3, column=col)
            cell.border = border
        
        # Row 4: Sub-headers
        ws['B4'] = "ID function / ID technical approach"
        ws['B4'].font = header_font
        ws['B4'].alignment = center_alignment
        ws['B4'].fill = derivation_fill
        ws['B4'].border = border
        
        ws['C4'] = "Asset candidate"
        ws['C4'].font = header_font
        ws['C4'].alignment = center_alignment
        ws['C4'].fill = derivation_fill
        ws['C4'].border = border
        
        ws['D4'] = "C – Confidentiality"
        ws['D4'].font = header_font
        ws['D4'].alignment = center_alignment
        ws['D4'].fill = security_fill
        ws['D4'].border = border
        
        ws['E4'] = "I – Integrity"
        ws['E4'].font = header_font
        ws['E4'].alignment = center_alignment
        ws['E4'].fill = security_fill
        ws['E4'].border = border
        
        ws['F4'] = "A - Availability"
        ws['F4'].font = header_font
        ws['F4'].alignment = center_alignment
        ws['F4'].fill = security_fill
        ws['F4'].border = border
        
        ws['G4'] = "The asset candidate requires further analysis?"
        ws['G4'].font = header_font
        ws['G4'].alignment = center_alignment
        ws['G4'].fill = cyber_assets_fill
        ws['G4'].border = border
        
        ws['H4'] = "Identified asset ID"
        ws['H4'].font = header_font
        ws['H4'].alignment = center_alignment
        ws['H4'].fill = cyber_assets_fill
        ws['H4'].border = border
        
        # Row 5: Functional approach (B-H merged)
        ws.merge_cells('B5:H5')
        ws['B5'] = "Functional approach"
        ws['B5'].font = header_font
        ws['B5'].alignment = center_alignment
        ws['B5'].fill = approach_fill
        
        for col in range(2, 9):
            cell = ws.cell(row=5, column=col)
            cell.border = border
        
        # Rows 6-15: FA01-FA10 data rows
        for i, row in enumerate(range(6, 16), 1):
            # Column B: FA labels with derivation background
            ws[f'B{row}'] = f"FA{i:02d}"
            ws[f'B{row}'].fill = derivation_fill
            ws[f'B{row}'].border = border
            
            # Column C: derivation background
            ws[f'C{row}'].fill = derivation_fill
            ws[f'C{row}'].border = border
            
            # Columns D, E, F, G: data background
            for col in ['D', 'E', 'F', 'G']:
                ws[f'{col}{row}'].fill = data_fill
                ws[f'{col}{row}'].border = border
            
            # Column H: cyber assets background
            ws[f'H{row}'].fill = cyber_assets_fill
            ws[f'H{row}'].border = border
        
        # Row 16: Technical approach (B-H merged)
        ws.merge_cells('B16:H16')
        ws['B16'] = "Technical approach"
        ws['B16'].font = header_font
        ws['B16'].alignment = center_alignment
        ws['B16'].fill = approach_fill
        
        for col in range(2, 9):
            cell = ws.cell(row=16, column=col)
            cell.border = border
        
        # Rows 17-26: TA01-TA10 data rows
        for i, row in enumerate(range(17, 27), 1):
            # Column B: TA labels with derivation background
            ws[f'B{row}'] = f"TA{i:02d}"
            ws[f'B{row}'].fill = derivation_fill
            ws[f'B{row}'].border = border
            
            # Column C: derivation background
            ws[f'C{row}'].fill = derivation_fill
            ws[f'C{row}'].border = border
            
            # Columns D, E, F, G: data background
            for col in ['D', 'E', 'F', 'G']:
                ws[f'{col}{row}'].fill = data_fill
                ws[f'{col}{row}'].border = border
            
            # Column H: cyber assets background
            ws[f'H{row}'].fill = cyber_assets_fill
            ws[f'H{row}'].border = border
    
    def _format_tara_sheet(self, ws):
        """Format the TARA sheet with complex column structure and styling"""
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Define all colors
        main_header_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        assets_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
        damage_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        impact_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cyber_control_fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
        stride_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        attack_path_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        potential_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
        risk_eval_fill = PatternFill(start_color="D5E8D4", end_color="D5E8D4", fill_type="solid")
        risk_treatment_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        even_row_fill = PatternFill(start_color="D0CECE", end_color="D0CECE", fill_type="solid")
        
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        title_font = Font(name='Calibri', size=16, color='000000')
        header_font = Font(name='Calibri', size=11, color='000000')
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Set all column widths as specified
        column_widths = {
            'A': 8.58, 'B': 12.68, 'C': 23.26, 'D': 14.68, 'E': 14.58, 'F': 35, 'G': 15.26, 
            'H': 4.42, 'I': 15.26, 'J': 4.42, 'K': 15.26, 'L': 4.42, 'M': 15.26, 'N': 4.42, 
            'O': 4.42, 'P': 15.26, 'Q': 39.58, 'R': 10.26, 'S': 30, 'T': 37.11, 'U': 37.11, 
            'V': 37.11, 'W': 37.11, 'X': 37.11, 'Y': 37.11, 'Z': 62.11, 'AA': 15.26, 'AB': 4.42, 
            'AC': 15.26, 'AD': 4.42, 'AE': 15.26, 'AF': 4.42, 'AG': 15.26, 'AH': 4.42, 'AI': 15.26, 
            'AJ': 4.42, 'AK': 15.26, 'AL': 15.26, 'AM': 15.26, 'AN': 15.26, 'AO': 28.26, 'AP': 28.26, 
            'AQ': 28.26, 'AR': 28.26, 'AS': 28.26, 'AT': 28.26, 'AU': 8.58
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Set row heights as specified
        ws.row_dimensions[1].height = 36
        ws.row_dimensions[3].height = 23.3
        ws.row_dimensions[4].height = 21.8
        
        # Odd rows 7-51 have height 57.8
        for row in range(7, 52, 2):  # Every odd row from 7 to 51
            ws.row_dimensions[row].height = 57.8
        
        # Line 1: Main title B-AT merged
        ws.merge_cells('B1:AT1')
        ws['B1'] = "Threat analysis and risk assessment"
        ws['B1'].font = title_font
        ws['B1'].alignment = center_alignment
        ws['B1'].fill = main_header_fill
        
        # Apply borders to main title
        for col_num in range(2, 47):  # B to AT
            cell = ws.cell(row=1, column=col_num)
            cell.border = border
        
        # Define specific TARA section colors and bold border
        threat_scenario_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        attack_feasibility_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
        cyber_goals_fill = PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid")
        
        bold_border = Border(
            left=Side(style='thick', color='000000'),
            right=Side(style='thick', color='000000'),
            top=Side(style='thick', color='000000'),
            bottom=Side(style='thick', color='000000')
        )
        
        # Line 3: Section headers with proper merging and colors
        section_headers = [
            ('B3:E3', 'Assets', assets_fill),
            ('F3:Q3', 'Impact analysis', impact_fill),
            ('R3:S3', 'Cybersecurity control', cyber_control_fill),
            ('T3:Y3', 'Threat scenario', threat_scenario_fill),
            ('AA3:AN3', 'Attack feasibility assessment', attack_feasibility_fill),
            ('AO3:AT3', 'Cybersecurity goals and claims', cyber_goals_fill)
        ]
        

        
        for cell_range, text, fill in section_headers:
            ws.merge_cells(cell_range)
            start_cell = cell_range.split(':')[0]
            ws[start_cell] = text
            ws[start_cell].font = header_font
            ws[start_cell].alignment = center_alignment
            ws[start_cell].fill = fill
        
        # Z: Attack Path (merged 3-5)
        ws.merge_cells('Z3:Z5')
        ws['Z3'] = "Attack Path"
        ws['Z3'].font = header_font
        ws['Z3'].alignment = center_alignment
        ws['Z3'].fill = attack_path_fill
        
        # Apply borders to line 3
        for col_num in range(2, 47):  # B to AT
            cell = ws.cell(row=3, column=col_num)
            cell.border = border
        
        # Lines 4-5: Detailed headers with proper merging and corrected titles
        # Asset headers (B-E as specified)
        asset_headers = [
            ('B4:B5', 'Asset ID'),
            ('C4:C5', 'Asset'),
            ('D4:D5', 'Security property loss'),
            ('E4:E5', 'Stakeholder')
        ]
        
        for cell_range, text in asset_headers:
            ws.merge_cells(cell_range)
            start_cell = cell_range.split(':')[0]
            ws[start_cell] = text
            ws[start_cell].font = header_font
            ws[start_cell].alignment = center_alignment
            if start_cell[0] in ['B', 'C', 'D']:
                ws[start_cell].fill = assets_fill
            elif start_cell[0] in ['E']:
                ws[start_cell].fill = damage_fill
        
        # Impact section headers
        ws.merge_cells('G4:P4')
        ws['G4'] = "Impact category"
        ws['G4'].font = header_font
        ws['G4'].alignment = center_alignment
        ws['G4'].fill = impact_fill
        
        ws.merge_cells('Q4:Q5')
        ws['Q4'] = "Justification(optional)"
        ws['Q4'].font = header_font
        ws['Q4'].alignment = center_alignment
        ws['Q4'].fill = impact_fill
        
        # Impact categories (line 5)
        impact_categories = [
            ('G5:H5', 'Safety'),
            ('I5:J5', 'Financial'),
            ('K5:L5', 'Operational'),
            ('M5:N5', 'Privacy'),
            ('O5:P5', 'Impact Level')
        ]
        
        for cell_range, text in impact_categories:
            ws.merge_cells(cell_range)
            start_cell = cell_range.split(':')[0]
            ws[start_cell] = text
            ws[start_cell].font = header_font
            ws[start_cell].alignment = center_alignment
            ws[start_cell].fill = impact_fill
        
        # Apply borders to all header rows
        for row in [4, 5]:
            for col_num in range(1, 47):  # A to AT
                cell = ws.cell(row=row, column=col_num)
                cell.border = border
        
        # Complete remaining headers for TARA sheet
        
        # Cybersecurity control headers (R-S)
        ws.merge_cells('R4:R5')
        ws['R4'] = "ID"
        ws['R4'].font = header_font
        ws['R4'].alignment = center_alignment
        ws['R4'].fill = cyber_control_fill
        
        ws.merge_cells('S4:S5')
        ws['S4'] = "Cybersecurity control"
        ws['S4'].font = header_font
        ws['S4'].alignment = center_alignment
        ws['S4'].fill = cyber_control_fill
        
        # STRIDE headers (T-Y)
        ws.merge_cells('T4:Y4')
        ws['T4'] = "Attack method – STRIDE"
        ws['T4'].font = header_font
        ws['T4'].alignment = center_alignment
        ws['T4'].fill = stride_fill
        
        stride_categories = ['Spoofing', 'Tampering', 'Repudiation', 'Information disclosure', 'Denial of service', 'Elevation of privileges']
        for i, category in enumerate(stride_categories):
            col_letter = get_column_letter(20 + i)  # T=20, U=21, etc.
            ws[f'{col_letter}5'] = category
            ws[f'{col_letter}5'].font = header_font
            ws[f'{col_letter}5'].alignment = center_alignment
            ws[f'{col_letter}5'].fill = stride_fill
        
        # Attack potential evaluation headers (AA-AK)
        ws.merge_cells('AA4:AK4')
        ws['AA4'] = "Attack potential evaluation"
        ws['AA4'].font = header_font
        ws['AA4'].alignment = center_alignment
        ws['AA4'].fill = potential_fill
        
        attack_potential_categories = [
            ('AA5:AB5', 'Elapsed time'),
            ('AC5:AD5', 'Specialist expertise'),
            ('AE5:AF5', 'Knowledge of the item or component'),
            ('AG5:AH5', 'Window of opportunity'),
            ('AI5:AJ5', 'Equipment'),
            ('AK5:AK5', 'Attack Vector')
        ]
        
        for cell_range, text in attack_potential_categories:
            ws.merge_cells(cell_range)
            start_cell = cell_range.split(':')[0]
            ws[start_cell] = text
            ws[start_cell].font = header_font
            ws[start_cell].alignment = center_alignment
            ws[start_cell].fill = potential_fill
        
        # Risk evaluation headers (AL-AN)
        ws.merge_cells('AL4:AN4')
        ws['AL4'] = "Risk evaluation"
        ws['AL4'].font = header_font
        ws['AL4'].alignment = center_alignment
        ws['AL4'].fill = risk_eval_fill
        
        risk_eval_categories = ['Summary', 'Attack feasibility', 'Risk determination']
        for i, category in enumerate(risk_eval_categories):
            col_letter = get_column_letter(38 + i)  # AL=38, AM=39, AN=40
            ws[f'{col_letter}5'] = category
            ws[f'{col_letter}5'].font = header_font
            ws[f'{col_letter}5'].alignment = center_alignment
            ws[f'{col_letter}5'].fill = risk_eval_fill
        
        # Risk treatment headers (AO-AT)
        risk_treatment_categories = [
            ('AO4:AO5', 'Risk threshold level'),
            ('AP4:AP5', 'Risk treatment option'),
            ('AQ4:AQ5', 'ID'),
            ('AR4:AR5', 'Cybersecurity goal'),
            ('AS4:AS5', 'Cyber security claim'),
            ('AT4:AT5', 'CAL')
        ]
        
        for cell_range, text in risk_treatment_categories:
            ws.merge_cells(cell_range)
            start_cell = cell_range.split(':')[0]
            ws[start_cell] = text
            ws[start_cell].font = header_font
            ws[start_cell].alignment = center_alignment
            ws[start_cell].fill = risk_treatment_fill
        
        # Apply background colors and borders to data area (rows 6-52)
        for row in range(6, 53):
            # Even rows: bold black outline
            if row % 2 == 0:
                for col_num in range(2, 47):  # B to AT
                    cell = ws.cell(row=row, column=col_num)
                    cell.border = bold_border
            else:
                # Odd rows: apply section background colors and bold borders
                for col_num in range(2, 47):  # B to AT
                    cell = ws.cell(row=row, column=col_num)
                    cell.border = bold_border
                    
                    col_letter = get_column_letter(col_num)
                    # T-Y: Threat scenario background (#FFF2CC)
                    if col_letter in ['T', 'U', 'V', 'W', 'X', 'Y']:
                        cell.fill = threat_scenario_fill
                    # AA-AN: Attack feasibility assessment background (#C6E0B4)
                    elif col_letter in ['AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN']:
                        cell.fill = attack_feasibility_fill
                    # AO-AT: Cybersecurity goals and claims background (#EDEDED)
                    elif col_letter in ['AO', 'AP', 'AQ', 'AR', 'AS', 'AT']:
                        cell.fill = cyber_goals_fill
        
        # Apply bold borders where color changes in the header
        for row in [3, 4, 5]:
            for col_num in range(2, 47):  # B to AT
                cell = ws.cell(row=row, column=col_num)
                cell.border = bold_border
        
        # Make header rows 1-5 frozen for scrolling
        ws.freeze_panes = 'A6'
    
    def _format_matrices_guidelines_sheet(self, ws):
        """Format the Matrices and guidelines sheet with exact specifications"""
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Define all colors exactly as specified
        asset_fill = PatternFill(start_color="9BC2E6", end_color="9BC2E6", fill_type="solid")
        security_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        tara_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        impact_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        risk_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
        goals_fill = PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid")
        
        # Risk value colors
        green_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        orange_fill = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        
        # Borders
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        thick_border = Border(
            left=Side(style='thick', color='000000'),
            right=Side(style='thick', color='000000'),
            top=Side(style='thick', color='000000'),
            bottom=Side(style='thick', color='000000')
        )
        
        # Fonts
        title_font = Font(name='Calibri', size=12, bold=True, color='000000')
        header_font = Font(name='Calibri', size=11, bold=True, color='000000')
        regular_font = Font(name='Calibri', size=11, color='000000')
        
        # Alignments
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        
        # Set column widths exactly as specified
        column_widths = {
            'A': 8.58, 'B': 47.42, 'C': 8.58, 'D': 17.42, 'E': 17.42, 'F': 8.58,
            'G': 18.84, 'H': 18.84, 'I': 18.84, 'J': 8.58, 'K': 35, 'L': 8.84,
            'M': 35, 'N': 8.84, 'O': 35, 'P': 8.84, 'Q': 35, 'R': 8.84, 'S': 35,
            'T': 8.84, 'U': 8.84, 'V': 8.84, 'W': 8.84, 'X': 8.84, 'Y': 8.84,
            'Z': 8.84, 'AA': 8.84, 'AB': 8.84, 'AC': 8.84, 'AD': 8.84, 'AE': 8.84
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Column B (Rows 2-5)
        ws['B2'] = "Asset identification process"
        ws['B2'].font = title_font
        ws['B2'].fill = asset_fill
        ws['B2'].alignment = center_align
        
        ws['B3'] = "Determination of security property loss"
        ws['B3'].font = header_font
        ws['B3'].fill = security_fill
        ws['B3'].alignment = center_align
        
        ws['B4'] = "Negligible impact"
        ws['B4'].font = regular_font
        ws['B4'].alignment = left_align
        
        ws['B5'] = "Non negligible impact"
        ws['B5'].font = regular_font
        ws['B5'].alignment = left_align
        
        # Apply borders to B2:B5 with thick outline on B2 and B3
        for row in range(2, 6):
            cell = ws[f'B{row}']
            if row in [2, 3]:
                cell.border = thick_border
            else:
                cell.border = thin_border
        
        # Row 2: Main TARA header (D2:AD2)
        ws.merge_cells('D2:AD2')
        ws['D2'] = "Threat analysis and risk assessment"
        ws['D2'].font = title_font
        ws['D2'].fill = tara_fill
        ws['D2'].alignment = center_align
        
        # Apply borders to merged header
        for col_num in range(4, 31):  # D to AD
            cell = ws.cell(row=2, column=col_num)
            cell.border = thin_border
        
        # Impact Analysis Section (D3:E8)
        ws.merge_cells('D3:E3')
        ws['D3'] = "Impact analysis - category values"
        ws['D3'].font = header_font
        ws['D3'].fill = impact_fill
        ws['D3'].alignment = center_align
        
        ws['D4'] = "Impact category"
        ws['D4'].font = header_font
        ws['D4'].fill = impact_fill
        ws['D4'].alignment = center_align
        
        ws['E4'] = "Impact value"
        ws['E4'].font = header_font
        ws['E4'].fill = impact_fill
        ws['E4'].alignment = center_align
        
        # Impact data
        impact_data = [
            ('Negligible', 0), ('Moderate', 1), ('Major', 3), ('Severe', 5)
        ]
        
        for i, (category, value) in enumerate(impact_data):
            row = 5 + i
            ws[f'D{row}'] = category
            ws[f'D{row}'].font = regular_font
            ws[f'D{row}'].alignment = left_align
            ws[f'D{row}'].border = thin_border
            
            ws[f'E{row}'] = value
            ws[f'E{row}'].font = regular_font
            ws[f'E{row}'].alignment = center_align
            ws[f'E{row}'].border = thin_border
            ws[f'E{row}'].number_format = '0'  # Numeric format
        
        # Apply borders to D3:E8 table
        for row in range(3, 9):
            for col in ['D', 'E']:
                ws[f'{col}{row}'].border = thick_border
        
        # Impact analysis – impact level estimation (G3:I8)
        ws.merge_cells('G3:I3')
        ws['G3'] = "Impact analysis – impact level estimation"
        ws['G3'].font = header_font
        ws['G3'].fill = impact_fill
        ws['G3'].alignment = center_align
        
        # Headers for G4:I4
        ws['G4'] = "Impact level"
        ws['G4'].font = header_font
        ws['G4'].fill = impact_fill
        ws['G4'].alignment = center_align
        
        ws['H4'] = "Values from"
        ws['H4'].font = header_font
        ws['H4'].fill = impact_fill
        ws['H4'].alignment = center_align
        
        ws['I4'] = "Values to"
        ws['I4'].font = header_font
        ws['I4'].fill = impact_fill
        ws['I4'].alignment = center_align
        
        # Impact level data G5:I8
        impact_levels = ['Negligible', 'Moderate', 'Major', 'Severe']
        values_from = [0, 3, 7, 11]
        values_to = [2, 6, 10, 1000]
        
        for i, (level, from_val, to_val) in enumerate(zip(impact_levels, values_from, values_to)):
            row = 5 + i
            ws[f'G{row}'] = level
            ws[f'G{row}'].font = regular_font
            ws[f'G{row}'].alignment = center_align
            
            ws[f'H{row}'] = from_val
            ws[f'H{row}'].font = regular_font
            ws[f'H{row}'].alignment = center_align
            ws[f'H{row}'].number_format = '0'
            
            ws[f'I{row}'] = to_val
            ws[f'I{row}'].font = regular_font
            ws[f'I{row}'].alignment = center_align
            ws[f'I{row}'].number_format = '0'
        
        # Apply borders to G3:I8 table
        for row in range(3, 9):
            for col in ['G', 'H', 'I']:
                ws[f'{col}{row}'].border = thick_border
        
        # Attack Potential Category Values (K3:T9)
        ws.merge_cells('K3:T3')
        ws['K3'] = "Attack potential – category values"
        ws['K3'].font = header_font
        ws['K3'].fill = risk_fill
        ws['K3'].alignment = center_align
        
        # Category headers row 4
        categories = [
            ('K4', 'Elapsed time'), ('L4', 'Value'),
            ('M4', 'Specialist expertise'), ('N4', 'Value'),
            ('O4', 'Knowledge of item'), ('P4', 'Value'),
            ('Q4', 'Window of opportunity'), ('R4', 'Value'),
            ('S4', 'Equipment'), ('T4', 'Value')
        ]
        
        for cell, text in categories:
            ws[cell] = text
            ws[cell].font = header_font
            ws[cell].alignment = center_align
        
        # Attack potential data rows 5-9
        elapsed_time = ['<=1 day', '<=1 week', '<=1 month', '<=6 months', '>6 months']
        elapsed_values = [0, 1, 4, 17, 19]
        
        expertise = ['Layman', 'Proficient', 'Expert', 'Multiple experts']
        expertise_values = [0, 3, 6, 8]
        
        knowledge = ['Public', 'Restricted', 'Confidential', 'Strictly confidential']
        knowledge_values = [0, 3, 7, 11]
        
        opportunity = ['Unlimited', 'Easy', 'Moderate', 'Difficult/None']
        opportunity_values = [0, 1, 4, 10]
        
        equipment = ['Standard', 'Specialized', 'Bespoke', 'Multiple bespoke']
        equipment_values = [0, 4, 7, 9]
        
        # Fill elapsed time data
        for i, (time, value) in enumerate(zip(elapsed_time, elapsed_values)):
            row = 5 + i
            ws[f'K{row}'] = time
            ws[f'K{row}'].font = regular_font
            ws[f'L{row}'] = value
            ws[f'L{row}'].font = regular_font
            ws[f'L{row}'].alignment = center_align
            ws[f'L{row}'].number_format = '0'
        
        # Fill other categories (only 4 rows for these)
        for i in range(4):
            row = 5 + i
            ws[f'M{row}'] = expertise[i]
            ws[f'M{row}'].font = regular_font
            ws[f'N{row}'] = expertise_values[i]
            ws[f'N{row}'].font = regular_font
            ws[f'N{row}'].alignment = center_align
            ws[f'N{row}'].number_format = '0'
            
            ws[f'O{row}'] = knowledge[i]
            ws[f'O{row}'].font = regular_font
            ws[f'P{row}'] = knowledge_values[i]
            ws[f'P{row}'].font = regular_font
            ws[f'P{row}'].alignment = center_align
            ws[f'P{row}'].number_format = '0'
            
            ws[f'Q{row}'] = opportunity[i]
            ws[f'Q{row}'].font = regular_font
            ws[f'R{row}'] = opportunity_values[i]
            ws[f'R{row}'].font = regular_font
            ws[f'R{row}'].alignment = center_align
            ws[f'R{row}'].number_format = '0'
            
            ws[f'S{row}'] = equipment[i]
            ws[f'S{row}'].font = regular_font
            ws[f'T{row}'] = equipment_values[i]
            ws[f'T{row}'].font = regular_font
            ws[f'T{row}'].alignment = center_align
            ws[f'T{row}'].number_format = '0'
        
        # Apply borders to each pair of columns K&L, M&N, O&P, Q&R, S&T
        for start_col in ['K', 'M', 'O', 'Q', 'S']:
            end_col = chr(ord(start_col) + 1)
            for row in range(3, 10):
                ws[f'{start_col}{row}'].border = thick_border
                ws[f'{end_col}{row}'].border = thick_border
        
        # Attack Potential & Risk Analysis Matrix (Rows 11-18)
        ws.merge_cells('D11:K11')
        ws['D11'] = "Threat analysis and risk assessment"
        ws['D11'].font = header_font
        ws['D11'].fill = tara_fill
        ws['D11'].alignment = center_align
        
        ws.merge_cells('D12:E12')
        ws['D12'] = "Attack potential – attack vector"
        ws['D12'].font = header_font
        ws['D12'].fill = risk_fill
        ws['D12'].alignment = center_align
        
        # Attack vectors with merged cells for each row
        vectors = ['Physical', 'Local', 'Adjacent', 'Network']
        for i, vector in enumerate(vectors):
            row = 13 + i
            ws.merge_cells(f'D{row}:E{row}')
            ws[f'D{row}'] = vector
            ws[f'D{row}'].font = regular_font
            ws[f'D{row}'].alignment = center_align
        
        # Apply black outline to D12:E16
        for row in range(12, 17):
            for col in ['D', 'E']:
                ws[f'{col}{row}'].border = thick_border
        
        # Attack feasibility section
        ws.merge_cells('G12:I12')
        ws['G12'] = "Attack feasibility"
        ws['G12'].font = header_font
        ws['G12'].fill = risk_fill
        ws['G12'].alignment = center_align
        
        # Headers for G13:I13
        ws['G13'] = "Impact level"
        ws['G13'].font = header_font
        ws['G13'].fill = risk_fill
        ws['G13'].alignment = center_align
        
        ws['H13'] = "Values from"
        ws['H13'].font = header_font
        ws['H13'].fill = risk_fill
        ws['H13'].alignment = center_align
        
        ws['I13'] = "Values to"
        ws['I13'].font = header_font
        ws['I13'].fill = risk_fill
        ws['I13'].alignment = center_align
        
        # Data for G14:I16 (shifted down from original)
        feasibility_levels = ['High', 'Medium', 'Low', 'Very low']
        feasibility_min = [0, 14, 20, 25]
        feasibility_max = [13, 19, 24, 1000]
        
        for i, (level, min_val, max_val) in enumerate(zip(feasibility_levels, feasibility_min, feasibility_max)):
            row = 14 + i
            ws[f'G{row}'] = level
            ws[f'G{row}'].font = regular_font
            ws[f'G{row}'].alignment = center_align
            ws[f'H{row}'] = min_val
            ws[f'H{row}'].font = regular_font
            ws[f'H{row}'].alignment = center_align
            ws[f'H{row}'].number_format = '0'
            ws[f'I{row}'] = max_val
            ws[f'I{row}'].font = regular_font
            ws[f'I{row}'].alignment = center_align
            ws[f'I{row}'].number_format = '0'
        
        # Apply borders to G12:I17
        for row in range(12, 18):
            for col in ['G', 'H', 'I']:
                ws[f'{col}{row}'].border = thick_border
        
        # Cybersecurity Goals (K12:K17)
        ws['K12'] = "Cybersecurity goals and claims"
        ws['K12'].font = header_font
        ws['K12'].fill = goals_fill
        ws['K12'].alignment = center_align
        
        ws['K13'] = "Risk treatment decision"
        ws['K13'].font = header_font
        ws['K13'].fill = goals_fill
        ws['K13'].alignment = center_align
        
        treatments = ['Avoid the risk', 'Reduce the risk', 'Share the risk', 'Retain the risk']
        for i, treatment in enumerate(treatments):
            row = 14 + i
            ws[f'K{row}'] = treatment
            ws[f'K{row}'].font = regular_font
            ws[f'K{row}'].alignment = center_align
        
        # Apply bold border around K12:K17
        for row in range(12, 18):
            ws[f'K{row}'].border = thick_border
        
        # Risk Evaluation Matrix (V3:AD9)
        ws.merge_cells('V3:AD3')
        ws['V3'] = "Risk evaluation"
        ws['V3'].font = header_font
        ws['V3'].fill = risk_fill
        ws['V3'].alignment = center_align
        
        ws.merge_cells('X4:AA4')
        ws['X4'] = "Attack feasibility rating"
        ws['X4'].font = header_font
        ws['X4'].alignment = center_align
        
        # Feasibility rating headers
        feasibility_headers = ['Very low', 'Low', 'Medium', 'High']
        for i, header in enumerate(feasibility_headers):
            col = chr(88 + i)  # X, Y, Z, AA
            if col == chr(91):  # After Z comes AA
                col = 'AA'
            ws[f'{col}5'] = header
            ws[f'{col}5'].font = header_font
            ws[f'{col}5'].alignment = center_align
        
        ws.merge_cells('V6:V9')
        ws['V6'] = "Impact rating"
        ws['V6'].font = header_font
        ws['V6'].fill = risk_fill
        ws['V6'].alignment = center_align
        
        # Impact levels
        impact_levels = ['Severe', 'Major', 'Moderate', 'Negligible']
        for i, level in enumerate(impact_levels):
            row = 6 + i
            ws[f'W{row}'] = level
            ws[f'W{row}'].font = header_font
            ws[f'W{row}'].fill = risk_fill
            ws[f'W{row}'].alignment = center_align
        
        # Risk matrix values with colors
        risk_matrix = [
            ['2', '3', '4', '5'],  # Severe
            ['1', '2', '3', '4'],  # Major
            ['1', '2', '2', '3'],  # Moderate
            ['1', '1', '1', '1']   # Negligible
        ]
        
        color_map = {'1': None, '2': green_fill, '3': yellow_fill, '4': orange_fill, '5': red_fill}
        
        for row_idx, risk_row in enumerate(risk_matrix):
            excel_row = 6 + row_idx
            for col_idx, value in enumerate(risk_row):
                if col_idx == 3:  # AA column
                    col = 'AA'
                else:
                    col = chr(88 + col_idx)  # X, Y, Z
                
                ws[f'{col}{excel_row}'] = value
                ws[f'{col}{excel_row}'].font = header_font
                ws[f'{col}{excel_row}'].alignment = center_align
                if color_map[value]:
                    ws[f'{col}{excel_row}'].fill = color_map[value]
        
        # Risk Value Mapping (AC4:AD9)
        ws.merge_cells('AC4:AD4')
        ws['AC4'] = "Risk value mapping"
        ws['AC4'].font = header_font
        ws['AC4'].fill = risk_fill
        ws['AC4'].alignment = center_align
        
        risk_values = [1, 2, 3, 4, 5]
        risk_labels = ['Very low', 'Low', 'Medium', 'High', 'Critical']
        
        for i, (value, label) in enumerate(zip(risk_values, risk_labels)):
            row = 5 + i
            ws[f'AC{row}'] = value
            ws[f'AC{row}'].font = regular_font
            ws[f'AC{row}'].alignment = right_align
            ws[f'AC{row}'].number_format = '0'
            
            ws[f'AD{row}'] = label
            ws[f'AD{row}'].font = header_font
            ws[f'AD{row}'].alignment = left_align
            if color_map[str(value)]:
                ws[f'AD{row}'].fill = color_map[str(value)]
        
        # Apply borders to AC4:AD9
        for row in range(4, 10):
            for col in ['AC', 'AD']:
                ws[f'{col}{row}'].border = thick_border
        
        # Additional comprehensive formatting for Matrices and Guidelines sheet
        # Define bold_font for matrices sheet
        bold_font = Font(name='Calibri', size=11, bold=True, color='000000')
        
        # Columns B rows 2-5 should be outlined in bold 
        for row in range(2, 6):
            ws[f'B{row}'].border = thick_border
        
        # Column D lines 5-8 should be center justified 
        for row in range(5, 9):
            ws[f'D{row}'].alignment = center_align
        
        # Columns K M O Q and S should be center justified 
        for col in ['K', 'M', 'O', 'Q', 'S']:
            for row in range(5, 10):
                ws[f'{col}{row}'].alignment = center_align
        
        # Replace <= with ≤ symbol in column K
        for row in range(5, 10):
            if ws[f'K{row}'].value and '<=' in str(ws[f'K{row}'].value):
                ws[f'K{row}'].value = str(ws[f'K{row}'].value).replace('<=', '≤')
        
        # Rows 4 and 5 columns V-AA should all have a background of #C6E0B4
        green_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
        for row in [4, 5]:
            for col_num in range(22, 27):  # V to AA
                col_letter = get_column_letter(col_num)
                ws[f'{col_letter}{row}'].fill = green_fill
                ws[f'{col_letter}{row}'].border = thick_border
        
        # All the headers and labels in columns V and W should have a bold black outline
        for row in [4, 5, 13, 14, 15, 16, 17, 18, 22, 23, 24, 25, 26, 27]:
            for col in ['V', 'W']:
                ws[f'{col}{row}'].border = thick_border
                ws[f'{col}{row}'].font = bold_font
                ws[f'{col}{row}'].alignment = center_align
        
        # Set values first, then merge cells for Threat analysis rows
        yellow_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        ws['V11'] = "Threat analysis and risk assessment"
        ws['V11'].font = bold_font
        ws['V11'].alignment = center_align
        ws['V11'].fill = yellow_fill
        ws['V11'].border = thick_border
        ws.merge_cells('V11:AA11')
        
        ws['V20'] = "Threat analysis and risk assessment"
        ws['V20'].font = bold_font
        ws['V20'].alignment = center_align
        ws['V20'].fill = yellow_fill
        ws['V20'].border = thick_border
        ws.merge_cells('V20:AA20')
        
        # Set values first, then merge cells for separate sections
        # Risk threshold level spans V13:W13 only (not to AA)
        ws['V13'] = "Risk threshold level"
        ws['V13'].font = bold_font
        ws['V13'].alignment = center_align
        ws['V13'].fill = green_fill
        ws['V13'].border = thick_border
        ws.merge_cells('V13:W13')
        
        # Attack feasibility rating spans X13:AA13
        ws['X13'] = "Attack feasibility rating"
        ws['X13'].font = bold_font
        ws['X13'].alignment = center_align
        ws['X13'].fill = green_fill
        ws['X13'].border = thick_border
        ws.merge_cells('X13:AA13')
        
        # Column V and W lines 14, 22 and 23 should be merged
        ws.merge_cells('V14:W14')
        ws.merge_cells('V22:W22')
        ws.merge_cells('V23:W23')
        
        # Line 14 column X should be labeled Very low, Y Low, Z Medium and AA High
        ws['X14'] = "Very low"
        ws['Y14'] = "Low"
        ws['Z14'] = "Medium"
        ws['AA14'] = "High"
        for col in ['X', 'Y', 'Z', 'AA']:
            ws[f'{col}14'].font = bold_font
            ws[f'{col}14'].alignment = center_align
            ws[f'{col}14'].fill = green_fill
            ws[f'{col}14'].border = thick_border
        
        # Column W line 15 should be labeled Severe 16 Major 17 Moderate and 18 Negligible
        ws['W15'] = "Severe"
        ws['W16'] = "Major"
        ws['W17'] = "Moderate"
        ws['W18'] = "Negligible"
        for row in [15, 16, 17, 18]:
            ws[f'W{row}'].font = bold_font
            ws[f'W{row}'].alignment = center_align
            ws[f'W{row}'].fill = green_fill
            ws[f'W{row}'].border = thick_border
        
        # Line 15-18 column X should have a background of #70AD47 and be labeled Below
        green_high = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        for row in [15, 16, 17, 18]:
            ws[f'X{row}'] = "Below"
            ws[f'X{row}'].font = bold_font
            ws[f'X{row}'].alignment = center_align
            ws[f'X{row}'].fill = green_high
            ws[f'X{row}'].border = thick_border
        
        # Row 18 columns Y, Z, and AA should also have a background of #70AD47 and be labeled Below
        for col in ['Y', 'Z', 'AA']:
            ws[f'{col}18'] = "Below"
            ws[f'{col}18'].font = bold_font
            ws[f'{col}18'].alignment = center_align
            ws[f'{col}18'].fill = green_high
            ws[f'{col}18'].border = thick_border
        
        # Columns Y-AA lines 15-17 should be labeled Above with a #FFC000 background
        orange_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        for row in [15, 16, 17]:
            for col in ['Y', 'Z', 'AA']:
                ws[f'{col}{row}'] = "Above"
                ws[f'{col}{row}'].font = bold_font
                ws[f'{col}{row}'].alignment = center_align
                ws[f'{col}{row}'].fill = orange_fill
                ws[f'{col}{row}'].border = thick_border
        
        # Set value first, then merge cells for CAL determination
        ws['V21'] = "CAL determination"
        ws['V21'].font = bold_font
        ws['V21'].alignment = center_align
        ws['V21'].fill = green_fill
        ws['V21'].border = thick_border
        ws.merge_cells('V21:AA21')
        
        # Set value first, then merge cells for Attack vector
        ws['X22'] = "Attack vector"
        ws['X22'].font = bold_font
        ws['X22'].alignment = center_align
        ws['X22'].fill = green_fill
        ws['X22'].border = thick_border
        ws.merge_cells('X22:AA22')
        
        # Line 23 attack vectors
        ws['X23'] = "Physical"
        ws['Y23'] = "Local"
        ws['Z23'] = "Adjacent"
        ws['AA23'] = "Network"
        for col in ['X', 'Y', 'Z', 'AA']:
            ws[f'{col}23'].font = bold_font
            ws[f'{col}23'].alignment = center_align
            ws[f'{col}23'].fill = green_fill
            ws[f'{col}23'].border = thick_border
        
        # Set value first, then merge cells for Impact rating
        ws['V24'] = "Impact rating"
        ws['V24'].font = bold_font
        ws['V24'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws['V24'].fill = green_fill
        ws['V24'].border = thick_border
        ws.merge_cells('V24:V27')
        
        # W column impact levels
        ws['W24'] = "Severe"
        ws['W25'] = "Major"
        ws['W26'] = "Moderate"
        ws['W27'] = "Negligible"
        for row in [24, 25, 26, 27]:
            ws[f'W{row}'].font = bold_font
            ws[f'W{row}'].alignment = center_align
            ws[f'W{row}'].fill = green_fill
            ws[f'W{row}'].border = thick_border
        
        # X24 Y25 and Z26 should all have the value 2 with a background color of #70AD47
        for cell in ['X24', 'Y25', 'Z26']:
            ws[cell] = 2
            ws[cell].font = bold_font
            ws[cell].alignment = center_align
            ws[cell].fill = green_high
            ws[cell].border = thick_border
            ws[cell].number_format = '0'
        
        # Y24, Z25 and AA26 should have a value of 3 and a background color of #FFC000
        for cell in ['Y24', 'Z25', 'AA26']:
            ws[cell] = 3
            ws[cell].font = bold_font
            ws[cell].alignment = center_align
            ws[cell].fill = orange_fill
            ws[cell].border = thick_border
            ws[cell].number_format = '0'
        
        # Z24 AA24 and AA25 should have a value of 4 and a background color of #FF0000
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        for cell in ['Z24', 'AA24', 'AA25']:
            ws[cell] = 4
            ws[cell].font = bold_font
            ws[cell].alignment = center_align
            ws[cell].fill = red_fill
            ws[cell].border = thick_border
            ws[cell].number_format = '0'
        
        # X25, X26 and Y26 should all have a value of 1
        for cell in ['X25', 'X26', 'Y26']:
            ws[cell] = 1
            ws[cell].font = bold_font
            ws[cell].alignment = center_align
            ws[cell].fill = green_high
            ws[cell].border = thick_border
            ws[cell].number_format = '0'
        
        # Line 27 columns X-AA should have a value of "-"
        for col in ['X', 'Y', 'Z', 'AA']:
            ws[f'{col}27'] = "-"
            ws[f'{col}27'].font = bold_font
            ws[f'{col}27'].alignment = center_align
            ws[f'{col}27'].fill = green_fill
            ws[f'{col}27'].border = thick_border
        
        # Lines 13 and 14 and 21, 22, and 23 columns V – AA should all have a background #C6E0B4
        for row in [13, 14, 21, 22, 23]:
            for col_num in range(22, 27):  # V to AA
                col_letter = get_column_letter(col_num)
                ws[f'{col_letter}{row}'].fill = green_fill
                ws[f'{col_letter}{row}'].font = bold_font
                ws[f'{col_letter}{row}'].alignment = center_align
                ws[f'{col_letter}{row}'].border = thick_border
        
        # In column V the text should be wrapped for Impact Rating
        ws['V24'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)