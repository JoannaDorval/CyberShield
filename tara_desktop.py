#!/usr/bin/env python3
"""
TARA Desktop Application
Threat Assessment and Remediation Analysis - Desktop Version

A comprehensive cybersecurity threat analysis application that processes threat models, 
system architecture diagrams, and cross-mapping data to generate professional TARA documents 
with MITRE ATT&CK framework integration.
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading
import os
import json
import logging
from datetime import datetime
from pathlib import Path
import sys
from typing import Dict, List, Any

# Import the streamlined processing modules - EMB3D focused only
from parsers import Embed3dJsonParser
from mitre_integration import MitreIntegrator
from mitre_embed import MitreEmbedIntegrator
# PDF generator removed - Excel reports only
from enhanced_excel_generator import EnhancedTaraExcelGenerator


class TaraDesktopApp:
    """Main desktop application class for TARA"""

    def __init__(self, root):
        self.root = root
        self.root.title("TARA - Threat Assessment and Remediation Analysis")
        self.root.geometry("1200x800")
        self.root.resizable(True, True)

        # Configure window to handle scrolling
        self.root.minsize(800, 600)

        # Set up logging
        self.setup_logging()

        # Initialize variables for EMB3D-only two-path approach
        self.json_heatmap_file = tk.StringVar()
        self.analysis_data = None

        # Two-path workflow: json_heatmap OR questionnaire
        self.workflow_mode = tk.StringVar(
            value="questionnaire")  # json_heatmap or questionnaire
        self.embed_properties = {
            'hardware': [],
            'system_software': [],
            'application_software': [],
            'networking': []
        }

        # Initialize EMB3D-focused processors only
        self.json_parser = Embed3dJsonParser()
        self.mitre_integrator = MitreIntegrator()
        self.embed_integrator = MitreEmbedIntegrator()
        # PDF report generator removed - Excel reports only
        self.excel_generator = EnhancedTaraExcelGenerator()

        # Create GUI
        self.create_widgets()

        self.logger.info("TARA Desktop Application initialized")

    def _on_mousewheel(self, event):
        """Handle mouse wheel scrolling - Windows/macOS"""
        self.main_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
    
    def _on_mousewheel_linux(self, event):
        """Handle mouse wheel scrolling - Linux"""
        if event.num == 4:
            self.main_canvas.yview_scroll(-1, "units")
        elif event.num == 5:
            self.main_canvas.yview_scroll(1, "units")

    def _on_canvas_configure(self, event):
        """Handle canvas resize to adjust scrollable frame width"""
        canvas_width = event.width
        self.main_canvas.itemconfig(self.canvas_window, width=canvas_width)

    def setup_logging(self):
        """Set up logging for the application"""
        # Create logs directory if it doesn't exist
        logs_dir = Path("logs")
        logs_dir.mkdir(exist_ok=True)

        # Configure logging
        log_filename = logs_dir / f"tara_desktop_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"

        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(log_filename),
                logging.StreamHandler(sys.stdout)
            ])

        self.logger = logging.getLogger(__name__)

    def create_widgets(self):
        """Create and arrange GUI widgets with scrollable interface"""
        # Configure root window
        self.root.configure(bg='#f0f0f0')

        # Create main canvas and scrollbar for scrollable interface
        self.main_canvas = tk.Canvas(self.root, bg='#f0f0f0')
        self.scrollbar = ttk.Scrollbar(self.root,
                                       orient="vertical",
                                       command=self.main_canvas.yview)
        self.scrollable_frame = ttk.Frame(self.main_canvas)

        # Configure scrollable frame
        self.scrollable_frame.bind(
            "<Configure>", lambda e: self.main_canvas.configure(
                scrollregion=self.main_canvas.bbox("all")))

        # Create window in canvas
        self.canvas_window = self.main_canvas.create_window(
            (0, 0), window=self.scrollable_frame, anchor="nw")

        # Configure canvas scrolling
        self.main_canvas.configure(yscrollcommand=self.scrollbar.set)

        # Pack canvas and scrollbar
        self.main_canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")

        # Bind mousewheel to canvas - enhanced for better cross-platform support
        self.main_canvas.bind("<MouseWheel>", self._on_mousewheel)
        self.scrollable_frame.bind("<MouseWheel>", self._on_mousewheel)
        # Bind mouse wheel for Linux/macOS compatibility
        self.main_canvas.bind("<Button-4>", self._on_mousewheel_linux)
        self.main_canvas.bind("<Button-5>", self._on_mousewheel_linux)
        self.scrollable_frame.bind("<Button-4>", self._on_mousewheel_linux)
        self.scrollable_frame.bind("<Button-5>", self._on_mousewheel_linux)

        # Main content frame
        main_frame = ttk.Frame(self.scrollable_frame, padding="20")
        main_frame.grid(row=0, column=0, sticky="nsew")

        # Configure grid weights
        self.scrollable_frame.columnconfigure(0, weight=1)
        self.scrollable_frame.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)

        # Bind canvas resize
        self.main_canvas.bind('<Configure>', self._on_canvas_configure)

        # Title
        title_label = ttk.Label(
            main_frame,
            text="TARA - Threat Assessment and Remediation Analysis",
            font=('Arial', 16, 'bold'))
        title_label.grid(row=0, column=0, columnspan=2, pady=(0, 20))

        # Configuration section
        self.create_configuration_section(main_frame)

        # File upload section (for EMB3D JSON heatmap)
        self.create_json_heatmap_upload_section(main_frame)

        # MITRE EMB3D questionnaire section
        self.create_embed_questionnaire_section(main_frame)

        # Progress and status section
        self.create_progress_section(main_frame)

        # Results section
        self.create_results_section(main_frame)

        # Buttons section
        self.create_buttons_section(main_frame)

    def create_configuration_section(self, parent):
        """Create two-path toggle configuration section"""
        config_frame = ttk.LabelFrame(parent,
                                      text="Analysis Method",
                                      padding="15")
        config_frame.grid(row=1,
                          column=0,
                          columnspan=2,
                          sticky="ew",
                          pady=(0, 20))
        config_frame.columnconfigure(1, weight=1)

        # Main instruction
        instruction_label = ttk.Label(config_frame,
                                      text="Choose your analysis method:",
                                      font=('Arial', 11, 'bold'))
        instruction_label.grid(row=0,
                               column=0,
                               columnspan=2,
                               sticky=tk.W,
                               pady=(0, 15))

        # Two-path workflow selection
        workflow_frame = ttk.Frame(config_frame)
        workflow_frame.grid(row=1,
                            column=0,
                            columnspan=2,
                            sticky="ew",
                            pady=(0, 10))
        workflow_frame.columnconfigure(0, weight=1)
        workflow_frame.columnconfigure(1, weight=1)

        # JSON Heatmap Upload Option
        json_heatmap_frame = ttk.LabelFrame(
            workflow_frame, text="Option 1: Upload EMB3D JSON Heatmap", padding="10")
        json_heatmap_frame.grid(row=0, column=0, sticky="ew", padx=(0, 10))

        ttk.Radiobutton(json_heatmap_frame,
                        text="Upload EMB3D JSON file",
                        variable=self.workflow_mode,
                        value="json_heatmap",
                        command=self.on_workflow_change).grid(row=0,
                                                              column=0,
                                                              sticky=tk.W,
                                                              pady=(0, 5))

        ttk.Label(
            json_heatmap_frame,
            text=
            "Import device properties from\nMITRE EMB3D website export",
            font=('Arial', 9),
            foreground='gray').grid(row=1, column=0, sticky=tk.W)

        # Questionnaire Option
        questionnaire_frame = ttk.LabelFrame(
            workflow_frame,
            text="Option 2: Complete Questionnaire",
            padding="10")
        questionnaire_frame.grid(row=0, column=1, sticky="ew", padx=(10, 0))

        ttk.Radiobutton(questionnaire_frame,
                        text="Complete MITRE EMB3D form",
                        variable=self.workflow_mode,
                        value="questionnaire",
                        command=self.on_workflow_change).grid(row=0,
                                                              column=0,
                                                              sticky=tk.W,
                                                              pady=(0, 5))

        ttk.Label(questionnaire_frame,
                  text="Answer guided questions about\nyour system components",
                  font=('Arial', 9),
                  foreground='gray').grid(row=1, column=0, sticky=tk.W)

    def create_json_heatmap_upload_section(self, parent):
        """Create EMB3D JSON heatmap file upload section"""
        self.upload_frame = ttk.LabelFrame(parent,
                                           text="EMB3D JSON Heatmap Upload",
                                           padding="15")
        self.upload_frame.grid(row=2,
                               column=0,
                               columnspan=2,
                               sticky="ew",
                               pady=(0, 20))
        self.upload_frame.columnconfigure(1, weight=1)

        # File selection
        ttk.Label(self.upload_frame,
                  text="Select EMB3D JSON File:",
                  font=('Arial', 10, 'bold')).grid(row=0,
                                                   column=0,
                                                   sticky=tk.W,
                                                   pady=(0, 5))

        file_frame = ttk.Frame(self.upload_frame)
        file_frame.grid(row=1,
                        column=0,
                        columnspan=2,
                        sticky="ew",
                        pady=(0, 10))
        file_frame.columnconfigure(0, weight=1)

        self.file_path_var = tk.StringVar()
        self.file_entry = ttk.Entry(file_frame,
                                    textvariable=self.file_path_var,
                                    state="readonly",
                                    width=50)
        self.file_entry.grid(row=0, column=0, sticky="ew", padx=(0, 10))

        self.browse_button = ttk.Button(file_frame,
                                        text="Browse",
                                        command=self.browse_json_heatmap)
        self.browse_button.grid(row=0, column=1)

        # Demo buttons for easy testing
        demo_frame = ttk.Frame(self.upload_frame)
        demo_frame.grid(row=2, column=0, columnspan=2, sticky="ew", pady=10)
        
        ttk.Label(demo_frame, text="Quick Test:", font=('Arial', 9, 'bold')).grid(row=0, column=0, sticky="w")
        ttk.Button(demo_frame, text="Sample Camera", command=self.load_sample_camera).grid(row=0, column=1, padx=5)
        ttk.Button(demo_frame, text="Sample Sensor", command=self.load_sample_sensor).grid(row=0, column=2, padx=5)
        
        # File format info
        info_label = ttk.Label(
            self.upload_frame,
            text=
            "Supported formats: .json (EMB3D JSON heatmap exported from MITRE EMB3D website)",
            font=('Arial', 9),
            foreground='gray')
        info_label.grid(row=3, column=0, columnspan=2, sticky=tk.W)

        # Initially hidden
        self.upload_frame.grid_remove()

    def create_embed_questionnaire_section(self, parent):
        """Create MITRE EMBED device properties section"""
        self.embed_frame = ttk.LabelFrame(parent,
                                          text="MITRE EMBED Device Properties",
                                          padding="10")
        self.embed_frame.grid(row=3,
                              column=0,
                              columnspan=2,
                              sticky="ew",
                              pady=(0, 20))
        self.embed_frame.columnconfigure(0, weight=1)

        # Instructions and clear button
        header_frame = ttk.Frame(self.embed_frame)
        header_frame.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        header_frame.columnconfigure(0, weight=1)

        instruction_label = ttk.Label(
            header_frame,
            text="Select the device properties that apply to your system:",
            font=('Arial', 10, 'bold'))
        instruction_label.grid(row=0, column=0, sticky=tk.W)

        # Buttons frame for demo and clear
        buttons_frame = ttk.Frame(header_frame)
        buttons_frame.grid(row=0, column=1, sticky=tk.E)
        
        self.demo_button = ttk.Button(buttons_frame,
                                      text="Demo Data",
                                      command=self.load_demo_questionnaire)
        self.demo_button.grid(row=0, column=0, padx=(0, 5))
        
        self.clear_button = ttk.Button(buttons_frame,
                                       text="Clear All",
                                       command=self.clear_all_fields)
        self.clear_button.grid(row=0, column=1)

        # Create notebook for categories
        self.embed_notebook = ttk.Notebook(self.embed_frame)
        self.embed_notebook.grid(row=1, column=0, sticky="ew", pady=(0, 10))

        # Initially hidden
        self.embed_frame.grid_remove()

        # Get device properties from MITRE EMBED integrator
        device_props = self.embed_integrator.get_device_properties_form()

        # Create tabs for each category
        self.embed_checkboxes = {}
        for category, properties in device_props.items():
            tab_frame = ttk.Frame(self.embed_notebook)
            self.embed_notebook.add(tab_frame,
                                    text=category.replace('_', ' ').title())

            # Create scrollable frame
            canvas = tk.Canvas(tab_frame, height=200)
            scrollbar = ttk.Scrollbar(tab_frame,
                                      orient="vertical",
                                      command=canvas.yview)
            scrollable_frame = ttk.Frame(canvas)

            scrollable_frame.bind(
                "<Configure>",
                lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

            canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
            canvas.configure(yscrollcommand=scrollbar.set)

            canvas.grid(row=0, column=0, sticky="nsew")
            scrollbar.grid(row=0, column=1, sticky="ns")

            tab_frame.columnconfigure(0, weight=1)
            tab_frame.rowconfigure(0, weight=1)

            # Add checkboxes for properties
            self.embed_checkboxes[category] = {}
            row = 0
            for prop_id, description in properties.items():
                var = tk.BooleanVar()
                checkbox = ttk.Checkbutton(scrollable_frame,
                                           text=f"{prop_id}: {description}",
                                           variable=var)
                checkbox.grid(row=row, column=0, sticky="w", pady=2, padx=5)
                self.embed_checkboxes[category][prop_id] = var
                row += 1

    def on_workflow_change(self):
        """Handle workflow mode change for two-path toggle"""
        mode = self.workflow_mode.get()
        if mode == "json_heatmap":
            # Show JSON heatmap upload, hide questionnaire
            self.upload_frame.grid()
            self.embed_frame.grid_remove()
        else:  # questionnaire mode
            # Show questionnaire, hide JSON heatmap upload
            self.upload_frame.grid_remove()
            self.embed_frame.grid()

    def clear_all_fields(self):
        """Clear all questionnaire fields"""
        for category in self.embed_checkboxes:
            for prop_id in self.embed_checkboxes[category]:
                self.embed_checkboxes[category][prop_id].set(False)
    
    def load_demo_questionnaire(self):
        """Load demo questionnaire responses for testing"""
        # Clear all first
        self.clear_all_fields()
        
        # Set some demo responses
        demo_selections = {
            "Hardware Platform": ["HW.1", "HW.2"],
            "Software Platform": ["SW.3", "SW.5"],
            "Communication": ["C.1", "C.3"],
            "Data": ["DAT.1", "DAT.3"]
        }
        
        for category, properties in demo_selections.items():
            if category in self.embed_checkboxes:
                for prop_id in properties:
                    if prop_id in self.embed_checkboxes[category]:
                        self.embed_checkboxes[category][prop_id].set(True)
        
        self.log_message("Demo questionnaire responses loaded")

    def create_progress_section(self, parent):
        """Create progress and status widgets"""
        # Progress frame
        progress_frame = ttk.LabelFrame(parent, text="Status", padding="10")
        progress_frame.grid(row=4,
                            column=0,
                            columnspan=2,
                            sticky="ew",
                            pady=(0, 20))
        progress_frame.columnconfigure(0, weight=1)

        # Progress bar
        self.progress_bar = ttk.Progressbar(progress_frame,
                                            mode='indeterminate')
        self.progress_bar.grid(row=0, column=0, sticky="ew", pady=(0, 10))

        # Status label
        self.status_label = ttk.Label(progress_frame,
                                      text="Ready to process files...")
        self.status_label.grid(row=1, column=0, sticky=tk.W)

    def create_results_section(self, parent):
        """Create results display widgets"""
        # Results frame
        results_frame = ttk.LabelFrame(parent,
                                       text="Analysis Results",
                                       padding="10")
        results_frame.grid(row=5,
                           column=0,
                           columnspan=2,
                           sticky="nsew",
                           pady=(0, 20))
        results_frame.columnconfigure(0, weight=1)
        results_frame.rowconfigure(0, weight=1)

        # Results text area with scrollbar
        self.results_text = scrolledtext.ScrolledText(results_frame,
                                                      height=15,
                                                      wrap=tk.WORD,
                                                      state=tk.DISABLED,
                                                      font=('Consolas', 10))
        self.results_text.grid(row=0, column=0, sticky="nsew")

    def create_buttons_section(self, parent):
        """Create action buttons for TARA analysis and Excel saving"""
        # Buttons frame
        buttons_frame = ttk.Frame(parent)
        buttons_frame.grid(row=6, column=0, columnspan=2, pady=(0, 10))

        # Generate TARA Analysis button
        self.analyze_button = ttk.Button(buttons_frame,
                                         text="Generate TARA Analysis",
                                         command=self.start_analysis,
                                         style='Accent.TButton')
        self.analyze_button.pack(side=tk.LEFT, padx=(0, 10))

        # Save Excel Report button (disabled until analysis is complete)
        self.save_excel_button = ttk.Button(buttons_frame,
                                            text="Save Excel Report",
                                            command=self.save_excel_report,
                                            state=tk.DISABLED)
        self.save_excel_button.pack(side=tk.LEFT, padx=(0, 10))

        # Clear button
        self.clear_button = ttk.Button(buttons_frame,
                                       text="Clear All",
                                       command=self.clear_all)
        self.clear_button.pack(side=tk.LEFT)

    def generate_excel_template(self):
        """Generate and save TARA Excel template with proper project information format"""
        try:
            # Ask user where to save the template
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Save TARA Template As...",
                initialfile="TARA_Template.xlsx")

            if filename:
                # Create TARA template with proper structure
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.utils import get_column_letter

                wb = Workbook()
                ws = wb.active
                ws.title = "TARA Template"

                # Set column widths exactly as specified in your format
                column_widths = {
                    'A': 14.285,
                    'B': 8.0,
                    'C': 10.43,
                    'D': 14.285,
                    'E': 10.0,
                    'F': 10.285,
                    'G': 10.0,
                    'H': 11.57,
                    'I': 10.57,
                    'J': 12.57,
                    'K': 11.57,
                    'L': 11.0,
                    'M': 10.57,
                    'N': 9.57
                }
                for col, width in column_widths.items():
                    ws.column_dimensions[col].width = width

                # Add project information header (rows 3-4) exactly as specified
                project_data = {
                    3: [
                        {
                            'column': 'A',
                            'value': 'Project Name:',
                            'font': Font(bold=True),
                            'alignment': Alignment(horizontal='right')
                        },
                        {
                            'column': 'B',
                            'value': None,
                            'font': Font()
                        },
                        {
                            'column': 'C',
                            'value': 'Project Number:',
                            'font': Font(bold=True),
                            'alignment': Alignment(horizontal='right')
                        },
                        {
                            'column': 'D',
                            'value': None,
                            'font': Font()
                        },
                    ],
                    4: [
                        {
                            'column': 'A',
                            'value': 'Analyst Name:',
                            'font': Font(bold=True),
                            'alignment': Alignment(horizontal='right')
                        },
                        {
                            'column': 'B',
                            'value': None,
                            'font': Font()
                        },
                        {
                            'column': 'C',
                            'value': 'Date:',
                            'font': Font(bold=True),
                            'alignment': Alignment(horizontal='right')
                        },
                        {
                            'column': 'D',
                            'value': None,
                            'font': Font(),
                            'number_format': 'm/d/yyyy'
                        },
                    ]
                }

                # Apply project information formatting exactly as provided
                for row_num, cells in project_data.items():
                    for cell_info in cells:
                        col = cell_info['column']
                        cell = ws[f"{col}{row_num}"]
                        cell.value = cell_info['value']
                        cell.font = cell_info['font']
                        if 'alignment' in cell_info:
                            cell.alignment = cell_info['alignment']
                        if 'number_format' in cell_info:
                            cell.number_format = cell_info['number_format']
                        cell.border = Border()

                # Add section headers for TARA analysis starting from row 6
                section_row = 6

                # Add main TARA analysis header
                ws[f"A{section_row}"].value = "TARA Analysis - Asset Information"
                ws[f"A{section_row}"].font = Font(bold=True, size=14)
                ws[f"A{section_row}"].fill = PatternFill(start_color="D9E1F2",
                                                         end_color="D9E1F2",
                                                         fill_type="solid")

                # Add column headers for analysis (row 8)
                headers_row = section_row + 2
                analysis_headers = [
                    "Asset Name", "Asset Type", "Description", "Criticality",
                    "Network Exposure", "Data Sensitivity",
                    "Security Controls", "Dependencies", "Vulnerabilities",
                    "Threat Sources"
                ]

                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4472C4",
                                          end_color="4472C4",
                                          fill_type="solid")

                for col, header in enumerate(analysis_headers, 1):
                    cell = ws.cell(row=headers_row, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center",
                                               vertical="center")

                # Add sample data rows for demonstration
                sample_data = [[
                    "Web Server", "Server", "Main application server", "High",
                    "Internet-facing", "Medium", "Firewall, SSL/TLS",
                    "Database Server", "CVE-2023-XXXX", "External Attackers"
                ],
                               [
                                   "Database Server", "Database",
                                   "Customer data storage", "Critical",
                                   "Internal", "High",
                                   "Access Controls, Encryption", "Web Server",
                                   "SQL Injection", "Internal Threats"
                               ],
                               [
                                   "User Workstation", "Endpoint",
                                   "Employee workstation", "Medium",
                                   "Internal", "Low", "Antivirus, Updates",
                                   "Network Infrastructure", "Malware",
                                   "Insider Threats"
                               ]]

                # Add sample data to demonstrate the format
                for row_idx, row_data in enumerate(sample_data,
                                                   headers_row + 1):
                    for col_idx, value in enumerate(row_data, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)

                # Save the template
                wb.save(filename)

                self.update_status(f"TARA template saved: {filename}")
                self.log_message(
                    f"TARA Excel template generated and saved to: {filename}")

                # Show success message
                messagebox.showinfo(
                    "TARA Template Generated",
                    f"TARA Excel template has been saved to:\n{filename}\n\nThis template includes proper project information formatting and follows the official TARA structure."
                )

        except Exception as e:
            error_msg = f"Error generating TARA template: {str(e)}"
            self.update_status(error_msg)
            self.log_message(error_msg)
            messagebox.showerror("Error", error_msg)

    def browse_json_heatmap(self):
        """Open file dialog for EMB3D JSON heatmap file"""
        filename = filedialog.askopenfilename(
            title="Select EMB3D JSON Heatmap File",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")])
        if filename:
            self.file_path_var.set(filename)
            self.json_heatmap_file.set(filename)
            self.log_message(
                f"EMB3D JSON heatmap file selected: {Path(filename).name}")
    
    def load_sample_camera(self):
        """Load sample security camera JSON heatmap for testing"""
        sample_file = "sample_embed3d_heatmap.json"
        if os.path.exists(sample_file):
            self.file_path_var.set(sample_file)
            self.json_heatmap_file.set(sample_file)
            self.log_message("Loaded sample security camera EMB3D data")
        else:
            self.log_message("Sample camera file not found")
    
    def load_sample_sensor(self):
        """Load sample IoT sensor JSON heatmap for testing"""
        sample_file = "sample_embed3d_minimal.json"
        if os.path.exists(sample_file):
            self.file_path_var.set(sample_file)
            self.json_heatmap_file.set(sample_file)
            self.log_message("Loaded sample IoT sensor EMB3D data")
        else:
            self.log_message("Sample sensor file not found")

    def validate_files(self):
        """Validate that input is provided for the selected workflow"""
        workflow_mode = self.workflow_mode.get()

        if workflow_mode == "questionnaire":
            # For questionnaire mode, check if any EMBED properties are selected
            has_properties = False
            for category, checkboxes in self.embed_checkboxes.items():
                if any(var.get() for var in checkboxes.values()):
                    has_properties = True
                    break

            if not has_properties:
                messagebox.showerror(
                    "Validation Error",
                    "Please select at least one device property from the MITRE EMB3D questionnaire."
                )
                return False

        else:  # json_heatmap mode
            # Check JSON heatmap file
            if not self.json_heatmap_file.get():
                messagebox.showerror(
                    "Validation Error",
                    "Please select an EMB3D JSON heatmap file.")
                return False

            if not os.path.exists(self.json_heatmap_file.get()):
                messagebox.showerror(
                    "Validation Error",
                    "The selected JSON heatmap file does not exist.")
                return False

        return True

    def start_analysis(self):
        """Start the analysis process in a separate thread"""
        # Validate files first
        if not self.validate_files():
            return

        # Disable buttons and start progress
        self.analyze_button.config(state=tk.DISABLED)
        self.progress_bar.start()

        # Clear previous results
        self.results_text.config(state=tk.NORMAL)
        self.results_text.delete(1.0, tk.END)
        self.results_text.config(state=tk.DISABLED)

        # Start analysis in separate thread
        analysis_thread = threading.Thread(target=self.perform_analysis)
        analysis_thread.daemon = True
        analysis_thread.start()

    def perform_analysis(self):
        """Perform enhanced threat analysis with support for multiple input types"""
        try:
            self.update_status("Starting enhanced TARA analysis...")
            self.log_message("=== Starting Enhanced TARA Analysis ===")

            # Determine workflow mode
            workflow_mode = self.workflow_mode.get()

            if workflow_mode == "questionnaire":
                # MITRE EMBED Questionnaire Mode
                analysis_data = self._process_questionnaire_mode()
            else:
                # EMB3D JSON Heatmap Mode
                analysis_data = self._process_json_heatmap_mode()

            # Integrate with MITRE frameworks
            analysis_data = self._integrate_mitre_frameworks(analysis_data)

            # Generate comprehensive analysis
            self._finalize_analysis(analysis_data)

        except Exception as e:
            error_msg = f"Analysis failed: {str(e)}"
            self.logger.error(error_msg, exc_info=True)
            self.update_status("Analysis failed!")
            self.log_message(f"ERROR: {error_msg}")
            messagebox.showerror("Analysis Error", error_msg)
        finally:
            # Re-enable buttons and stop progress
            self.root.after(0, self._analysis_complete)

    def _process_questionnaire_mode(self) -> Dict[str, Any]:
        """Process MITRE EMBED questionnaire responses"""
        self.update_status("Processing MITRE EMBED questionnaire...")
        self.log_message("Processing device properties from questionnaire...")

        # Collect selected properties
        selected_properties = {}
        for category, checkboxes in self.embed_checkboxes.items():
            selected_properties[category] = [
                prop_id for prop_id, var in checkboxes.items() if var.get()
            ]

        # Generate assessment from properties
        embed_assessment = self.embed_integrator.assess_device_properties(
            selected_properties)

        # Extract generated threats and assets
        threats = embed_assessment.get('threat_vectors', [])
        assets = self._generate_assets_from_properties(selected_properties)

        analysis_data = {
            'threats': threats,
            'assets': assets,
            'data_flows': [],
            'risks': [],
            'mitigations': embed_assessment.get('recommended_controls', []),
            'embed_assessment': embed_assessment,
            'metadata': {
                'source':
                'questionnaire',
                'input_type':
                'mitre_embed_questionnaire',
                'properties_count':
                sum(len(props) for props in selected_properties.values())
            }
        }

        self.log_message(
            f"Generated {len(threats)} threats from {analysis_data['metadata']['properties_count']} device properties"
        )
        return analysis_data

    def _process_json_heatmap_mode(self) -> Dict[str, Any]:
        """Process EMB3D JSON heatmap file analysis"""
        try:
            self.update_status("Processing EMB3D JSON heatmap file...")
            self.log_message(
                f"Processing JSON heatmap: {Path(self.json_heatmap_file.get()).name}"
            )

            analysis_data = {
                'assets': [],
                'threats': [],
                'mitigations': [],
                'metadata': {
                    'analysis_type': 'embed3d_json_heatmap',
                    'input_file': Path(self.json_heatmap_file.get()).name,
                    'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                }
            }

            # Process JSON heatmap file
            try:
                json_data = self.json_parser.parse(self.json_heatmap_file.get())
                if json_data:
                    # Extract device properties from JSON
                    selected_properties = json_data.get('selected_properties', {})
                    
                    # Generate assessment from properties using EMB3D integrator
                    embed_assessment = self.embed_integrator.assess_device_properties(selected_properties)
                    
                    # Extract threats and assets from assessment
                    analysis_data['threats'] = embed_assessment.get('threat_vectors', [])
                    analysis_data['assets'] = self._generate_assets_from_properties(selected_properties)
                    analysis_data['mitigations'] = embed_assessment.get('recommended_controls', [])
                    analysis_data['embed_assessment'] = embed_assessment
                    
                    self.log_message(
                        f"Found {len(selected_properties)} property categories with {len(analysis_data['threats'])} threats"
                    )
                else:
                    self.log_message("No data extracted from JSON heatmap file")
            except Exception as e:
                self.log_message(f"Error parsing JSON heatmap: {str(e)}")
                raise

            # Ensure we have some data for analysis
            if not any(analysis_data['assets']) and not analysis_data['threats']:
                raise ValueError("No valid data found in JSON heatmap file")

            return analysis_data

        except Exception as e:
            self.log_message(f"Error in JSON heatmap processing: {str(e)}")
            raise

    def _integrate_mitre_frameworks(
            self, analysis_data: Dict[str, Any]) -> Dict[str, Any]:
        """Integrate analysis with MITRE frameworks - always includes both ATT&CK and EMBED"""

        # Always perform MITRE EMBED assessment for all assets
        self.update_status("Processing MITRE EMBED assessment...")

        # Auto-assess all assets against EMBED properties
        embed_assessment = self._auto_assess_assets_with_embed(
            analysis_data.get('assets', []))
        analysis_data['embed_assessment'] = embed_assessment

        # Add EMBED-derived threats to existing threats
        embed_threats = embed_assessment.get('threat_vectors', [])
        existing_threats = analysis_data.get('threats', [])
        analysis_data['threats'] = existing_threats + embed_threats

        # Add EMBED controls to mitigations
        embed_controls = embed_assessment.get('controls', [])
        existing_mitigations = analysis_data.get('mitigations', [])
        analysis_data['mitigations'] = existing_mitigations + embed_controls

        self.log_message(
            f"Auto-assessed {len(analysis_data.get('assets', []))} assets against MITRE EMBED"
        )
        self.log_message(
            f"Added {len(embed_threats)} EMBED threat vectors and {len(embed_controls)} controls"
        )

        # Always perform MITRE ATT&CK integration with all threats (original + EMBED-derived)
        self.update_status("Mapping to MITRE ATT&CK...")
        self.log_message("Integrating with MITRE ATT&CK framework...")

        # Load cross-mapping data if available, otherwise use empty dict
        # Use built-in MITRE ATT&CK mappings for streamlined workflow
        crossmap_data = {}
        self.log_message("Using built-in MITRE ATT&CK mappings")

        # Perform MITRE integration with all threats
        all_threats = analysis_data.get('threats', [])
        mitre_mappings = self.mitre_integrator.map_threats_to_mitre(
            all_threats, crossmap_data)
        analysis_data['mitre_mappings'] = mitre_mappings

        # Generate recommendations based on all threats and mitigations
        recommendations = self.mitre_integrator.generate_recommendations(
            all_threats, mitre_mappings, analysis_data.get('mitigations', []))
        analysis_data['recommendations'] = recommendations

        self.log_message(
            f"Mapped {len(mitre_mappings.get('technique_mappings', []))} MITRE ATT&CK techniques"
        )

        return analysis_data

    def _auto_assess_assets_with_embed(
            self, assets: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Automatically assess assets against MITRE EMBED device properties"""
        # Determine device properties based on asset types and characteristics
        inferred_properties = {
            'hardware': [],
            'system_software': [],
            'application_software': [],
            'networking': []
        }

        for asset in assets:
            asset_type = asset.get('type', '').lower()
            asset_name = asset.get('name', '').lower()
            asset_desc = asset.get('description', '').lower()

            # Infer hardware properties
            if any(
                    term in asset_type or term in asset_name
                    or term in asset_desc for term in
                ['server', 'device', 'sensor', 'iot', 'embedded', 'hardware']):
                inferred_properties['hardware'].extend([
                    'PID-11',  # Microprocessor
                    'PID-12',  # Memory/Storage
                    'PID-13'  # Firmware/BIOS
                ])

            # Infer system software properties
            if any(term in asset_type or term in asset_name
                   or term in asset_desc for term in
                   ['server', 'system', 'database', 'operating', 'os']):
                inferred_properties['system_software'].extend([
                    'PID-21',  # Bootloader
                    'PID-23',  # Operating system
                    'PID-24',  # Device drivers
                    'PID-25',  # System services
                    'PID-28'  # Update mechanisms
                ])

            # Infer application software properties
            if any(
                    term in asset_type or term in asset_name
                    or term in asset_desc for term in
                ['application', 'app', 'web', 'api', 'service', 'software']):
                inferred_properties['application_software'].extend([
                    'PID-31',  # Application software
                    'PID-311',  # Web/HTTP applications
                    'PID-317',  # Communication protocols
                    'PID-319'  # API interfaces
                ])

            # Infer networking properties
            if any(term in asset_type or term in asset_name
                   or term in asset_desc for term in [
                       'network', 'gateway', 'router', 'firewall',
                       'load balancer', 'api'
                   ]):
                inferred_properties['networking'].extend([
                    'PID-41',  # Remote network services
                    'PID-411',  # Services with sensitive data
                    'PID-413',  # Wired networking
                    'PID-414'  # Network security protocols
                ])

        # Remove duplicates
        for category in inferred_properties:
            inferred_properties[category] = list(
                set(inferred_properties[category]))

        # Perform EMBED assessment with inferred properties
        if any(inferred_properties.values()):
            return self.embed_integrator.assess_device_properties(
                inferred_properties)
        else:
            # Default assessment for generic assets
            default_properties = {
                'hardware': ['PID-11', 'PID-12'],
                'system_software': ['PID-23', 'PID-25'],
                'application_software': ['PID-31'],
                'networking': ['PID-41']
            }
            return self.embed_integrator.assess_device_properties(
                default_properties)

    def _finalize_analysis(self, analysis_data: Dict[str, Any]):
        """Finalize analysis and enable Excel saving"""
        self.update_status("Finalizing analysis...")
        self.log_message("Generating final analysis report...")

        # Store analysis data with proper ID
        analysis_data['id'] = datetime.now().strftime('%Y%m%d_%H%M%S')
        analysis_data['timestamp'] = datetime.now()
        self.analysis_data = analysis_data

        # Generate analysis summary
        summary = self._generate_analysis_summary(analysis_data)

        # Display results
        self.root.after(0, lambda: self._display_results(summary))

        # Enable Excel save button
        self.save_excel_button.config(state=tk.NORMAL)

        self.update_status(
            "Analysis complete! You can now save the Excel report.")
        self.log_message("=== TARA Analysis Complete ===")

    def save_excel_report(self):
        """Save the TARA report as Excel with user-selected location"""
        if not self.analysis_data:
            messagebox.showerror("Error", "No analysis data available to save")
            return

        # Get save location from user
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Save TARA Excel Report",
            initialfile=
            f"TARA_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

        if filename:
            try:
                self.update_status("Generating Excel report...")
                self.log_message("Generating Excel report...")

                # Collect EMBED properties if selected
                embed_assessment = None
                selected_properties = {}
                for category, checkboxes in self.embed_checkboxes.items():
                    selected_properties[category] = [
                        prop_id for prop_id, var in checkboxes.items()
                        if var.get()
                    ]

                if any(selected_properties.values()):
                    embed_assessment = self.embed_integrator.assess_device_properties(
                        selected_properties)

                # Determine input type based on workflow
                input_type = "threat_model" if self.workflow_mode.get(
                ) == "threat_model" else "questionnaire"

                # Generate Excel report
                excel_path = self.excel_generator.generate_excel_report(
                    self.analysis_data,
                    input_type,
                    "both",  # Always use comprehensive analysis
                    embed_assessment)

                # Copy to user-selected location
                import shutil
                shutil.copy2(excel_path, filename)

                self.update_status("Excel report saved successfully!")
                self.log_message(f"Excel report saved to: {filename}")

                # Show success message
                messagebox.showinfo(
                    "Report Saved",
                    f"TARA Excel report has been saved to:\n{filename}\n\nThe report contains comprehensive threat analysis with MITRE ATT&CK and EMBED framework mappings."
                )

            except Exception as e:
                error_msg = f"Failed to save Excel report: {str(e)}"
                self.update_status("Failed to save Excel report!")
                self.log_message(error_msg)
                messagebox.showerror("Save Error", error_msg)

    def _auto_save_excel_report(self):
        """Automatically save Excel report with timestamped filename"""
        try:
            if not self.analysis_data:
                self.log_message(
                    "Warning: No analysis data available for Excel report")
                return

            # Generate timestamped filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"TARA_Excel_Report_{timestamp}.xlsx"

            # Collect EMBED properties if selected
            embed_assessment = None
            selected_properties = {}
            for category, checkboxes in self.embed_checkboxes.items():
                selected_properties[category] = [
                    prop_id for prop_id, var in checkboxes.items()
                    if var.get()
                ]

            if any(selected_properties.values()):
                embed_assessment = self.embed_integrator.assess_device_properties(
                    selected_properties)

            # Determine input type based on workflow
            input_type = "threat_model" if self.workflow_mode.get(
            ) == "threat_model" else "questionnaire"

            # Generate Excel report
            excel_path = self.excel_generator.generate_excel_report(
                self.analysis_data,
                input_type,
                "both",  # Always use comprehensive analysis
                embed_assessment)

            # Copy to working directory with clean filename
            import shutil
            shutil.copy2(excel_path, filename)

            self.update_status(f"Excel report saved: {filename}")
            self.log_message(
                f"Excel report automatically saved as: {filename}")

            # Show success message to user
            messagebox.showinfo(
                "Report Generated",
                f"TARA Excel report has been saved as:\n{filename}\n\nThe report contains comprehensive threat analysis with MITRE ATT&CK and EMBED framework mappings."
            )

        except Exception as e:
            error_msg = f"Error saving Excel report: {str(e)}"
            self.update_status(error_msg)
            self.log_message(error_msg)
            messagebox.showerror("Error", error_msg)

    def _generate_assets_from_properties(
            self,
            selected_properties: Dict[str, List[str]]) -> List[Dict[str, Any]]:
        """Generate asset list from selected device properties"""
        assets = []

        # Create assets based on property categories
        for category, properties in selected_properties.items():
            if properties:
                asset = {
                    'name': f'{category.replace("_", " ").title()} Component',
                    'type': category,
                    'description':
                    f'Device component representing {category} properties',
                    'properties': properties,
                    'criticality': 'Medium'
                }
                assets.append(asset)

        return assets

    def _generate_analysis_summary(self, analysis_data: Dict[str, Any]) -> str:
        """Generate comprehensive analysis summary"""
        summary_lines = [
            "TARA ANALYSIS SUMMARY", "=" * 50, "",
            f"Analysis Type: {analysis_data.get('metadata', {}).get('input_type', 'Unknown')}",
            f"Source: {analysis_data.get('metadata', {}).get('source', 'Unknown')}",
            "", "RESULTS OVERVIEW:",
            f"• Threats Identified: {len(analysis_data.get('threats', []))}",
            f"• Assets Analyzed: {len(analysis_data.get('assets', []))}",
            f"• Data Flows: {len(analysis_data.get('data_flows', []))}",
            f"• Mitigations: {len(analysis_data.get('mitigations', []))}", ""
        ]

        # Add MITRE mappings info
        if 'mitre_mappings' in analysis_data:
            mappings = analysis_data['mitre_mappings']
            summary_lines.extend([
                "MITRE ATT&CK INTEGRATION:",
                f"• Technique Mappings: {len(mappings.get('technique_mappings', []))}",
                f"• Recommendations: {len(analysis_data.get('recommendations', []))}",
                ""
            ])

        # Add EMBED assessment info
        if 'embed_assessment' in analysis_data:
            embed = analysis_data['embed_assessment']
            summary_lines.extend([
                "MITRE EMBED ASSESSMENT:",
                f"• Device Properties: {len(embed.get('security_implications', {}))}",
                f"• Threat Vectors: {len(embed.get('threat_vectors', []))}",
                f"• Controls: {len(embed.get('recommended_controls', []))}", ""
            ])

        # Add threat breakdown
        threats = analysis_data.get('threats', [])
        if threats:
            threat_categories = {}
            for threat in threats:
                category = threat.get('category', 'Unknown')
                threat_categories[category] = threat_categories.get(
                    category, 0) + 1

            summary_lines.extend([
                "THREAT CATEGORIES:", *[
                    f"• {category}: {count}"
                    for category, count in threat_categories.items()
                ], ""
            ])

        summary_lines.extend([
            "Analysis completed successfully.",
            "Use 'Save PDF Report' or 'Save Excel Report' to export results."
        ])

        return "\n".join(summary_lines)

    def _display_results(self, summary: str):
        """Display analysis results in the text widget"""
        self.results_text.config(state=tk.NORMAL)
        self.results_text.delete(1.0, tk.END)
        self.results_text.insert(tk.END, summary)
        self.results_text.config(state=tk.DISABLED)

        # Scroll to top
        self.results_text.see(1.0)

    def _analysis_complete(self):
        """Re-enable UI elements after analysis completion"""
        self.progress_bar.stop()
        self.analyze_button.config(state=tk.NORMAL)

    # PDF functionality removed - Excel reports only

    def save_excel_report(self):
        """Save the TARA report as Excel"""
        if not self.analysis_data:
            messagebox.showerror("Error", "No analysis data available to save")
            return

        # Get save location
        filename = filedialog.asksaveasfilename(
            title="Save TARA Excel Report",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=f"TARA_Excel_Report_{self.analysis_data['id']}.xlsx")

        if not filename:
            return

        try:
            self.update_status("Generating Excel report...")
            self.log_message("Generating Excel report...")

            # Collect EMBED properties if selected
            embed_assessment = None
            selected_properties = {}
            for category, checkboxes in self.embed_checkboxes.items():
                selected_properties[category] = [
                    prop_id for prop_id, var in checkboxes.items()
                    if var.get()
                ]

            if any(selected_properties.values()):
                embed_assessment = self.embed_integrator.assess_device_properties(
                    selected_properties)

            # Determine input type based on workflow
            input_type = "threat_model" if self.workflow_mode.get(
            ) == "threat_model" else "questionnaire"

            # Generate Excel report
            excel_path = self.excel_generator.generate_excel_report(
                self.analysis_data,
                input_type,
                "both",  # Always use comprehensive analysis
                embed_assessment)

            # Copy to desired location
            import shutil
            shutil.copy2(excel_path, filename)

            # Clean up temporary file
            if os.path.exists(excel_path):
                os.remove(excel_path)

            self.update_status("Excel report saved successfully!")
            self.log_message(f"Excel report saved to: {filename}")
            messagebox.showinfo(
                "Success",
                f"TARA Excel report saved successfully to:\n{filename}")

        except Exception as e:
            error_msg = f"Failed to save Excel report: {str(e)}"
            self.logger.error(error_msg, exc_info=True)
            self.update_status("Failed to save Excel report!")
            messagebox.showerror("Save Error", error_msg)

    def clear_all(self):
        """Clear all inputs and results"""
        # Clear file path
        self.threat_model_file.set("")
        if hasattr(self, 'file_path_var'):
            self.file_path_var.set("")

        # Clear questionnaire selections
        for category in self.embed_checkboxes:
            for prop_id in self.embed_checkboxes[category]:
                self.embed_checkboxes[category][prop_id].set(False)

        # Clear analysis data and results
        self.analysis_data = None

        self.results_text.config(state=tk.NORMAL)
        self.results_text.delete(1.0, tk.END)
        self.results_text.config(state=tk.DISABLED)

        # Reset button states
        self.save_excel_button.config(state=tk.DISABLED)

        self.update_status("Ready for analysis...")
        self.log_message("All fields cleared")

    def update_status(self, message):
        """Update status label (thread-safe)"""
        self.root.after(0, lambda: self.status_label.config(text=message))

    def log_message(self, message):
        """Log message to both logger and results area"""
        self.logger.info(message)
        # Also display in results if it's available
        if hasattr(self, 'results_text'):
            self.root.after(
                0, lambda: self._append_to_results(
                    f"[{datetime.now().strftime('%H:%M:%S')}] {message}"))

    def _append_to_results(self, message):
        """Append message to results text area (main thread only)"""
        self.results_text.config(state=tk.NORMAL)
        self.results_text.insert(tk.END, message + "\n")
        self.results_text.see(tk.END)
        self.results_text.config(state=tk.DISABLED)


def main():
    """Main function to run the application"""
    # Create the main window
    root = tk.Tk()

    # Set up the application style
    style = ttk.Style()
    try:
        # Try to use a modern theme if available
        style.theme_use('clam')
    except:
        # Fall back to default theme
        pass

    # Create and run the application
    app = TaraDesktopApp(root)

    # Center the window on screen
    root.update_idletasks()
    width = root.winfo_width()
    height = root.winfo_height()
    x = (root.winfo_screenwidth() // 2) - (width // 2)
    y = (root.winfo_screenheight() // 2) - (height // 2)
    root.geometry(f"{width}x{height}+{x}+{y}")

    # Start the GUI event loop
    root.mainloop()


if __name__ == "__main__":
    main()
