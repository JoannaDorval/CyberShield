"""
Enhanced Excel Report Generator for TARA Analysis with Precise Formatting
"""

import pandas as pd
import os
import logging
from typing import Dict, List, Any, Optional
from datetime import datetime

class EnhancedTaraExcelGenerator:
    """Generate comprehensive TARA Excel reports with multiple worksheets and precise formatting"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
    
    def generate_excel_report(self, analysis_data: Dict[str, Any], 
                            input_type: str, 
                            cross_ref_source: str,
                            embed_assessment: Optional[Dict] = None) -> str:
        """Generate complete Excel report with all analysis data and precise formatting"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"TARA_Excel_Report_{timestamp}.xlsx"
            
            # Create uploads directory if it doesn't exist
            output_dir = 'uploads'
            os.makedirs(output_dir, exist_ok=True)
            
            filepath = os.path.join(output_dir, filename)
            
            # Create writer object
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Create all required worksheets with precise formatting
                self._create_all_worksheets(writer.book, analysis_data, embed_assessment)
            
            self.logger.info(f"Excel report generated: {filepath}")
            return filepath
            
        except Exception as e:
            self.logger.error(f"Failed to generate Excel report: {e}")
            raise
    
    def _create_all_worksheets(self, workbook, analysis_data, embed_assessment):
        """Create all required worksheets with precise formatting"""
        # Remove default sheet if it exists
        if 'Sheet' in workbook.sheetnames:
            workbook.remove(workbook['Sheet'])
        
        # Create all worksheets with precise formatting as specified
        self._create_general_information_sheet(workbook, analysis_data)
        self._create_item_definition_sheet(workbook, analysis_data)
        self._create_item_definition_threat_model_sheet(workbook, analysis_data)
        self._create_asset_identification_sheet(workbook, analysis_data)
        self._create_tara_sheet(workbook, analysis_data)
        self._create_matrices_guidelines_sheet(workbook, analysis_data)
    
    def _create_general_information_sheet(self, workbook, analysis_data):
        """Create General Information sheet with precise formatting"""
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        ws = workbook.create_sheet('General Information')
        
        # Set precise row heights as specified
        ws.row_dimensions[1].height = 49.5
        ws.row_dimensions[2].height = 29.3
        ws.row_dimensions[3].height = 29.3
        for row in range(6, 20):
            ws.row_dimensions[row].height = 22.5
        
        # Define styles
        bold_font = Font(name='Calibri', size=11, bold=True)
        normal_font = Font(name='Calibri', size=11)
        bold_border = Border(
            left=Side(style='thick', color='000000'),
            right=Side(style='thick', color='000000'),
            top=Side(style='thick', color='000000'),
            bottom=Side(style='thick', color='000000')
        )
        
        # Header section with bold outline
        ws['A1'] = 'General Information'
        ws['A1'].font = bold_font
        ws['A1'].border = bold_border
        
        # Table headers (B-H, rows 6-19 with bold outline)
        table_headers = ['Field', 'Value', 'Description', 'Category', 'Status', 'Notes', 'Validation']
        for col_idx, header in enumerate(table_headers, start=2):
            cell = ws.cell(row=6, column=col_idx, value=header)
            cell.font = bold_font
            cell.border = bold_border
        
        # Add table data with bold outline for B-H columns, rows 6-19
        for row in range(7, 20):
            for col in range(2, 9):  # Columns B-H
                cell = ws.cell(row=row, column=col)
                cell.border = bold_border
        
        # Sheet ends at column I and after row 20 (rest unlined)
        ws.sheet_properties.tabColor = "1F4E79"
    
    def _create_item_definition_sheet(self, workbook, analysis_data):
        """Create Item Definition sheet with precise formatting"""
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        ws = workbook.create_sheet('Item Definition')
        
        # Set precise row heights as specified
        for row in range(4, 26):
            ws.row_dimensions[row].height = 23.3
        for row in range(27, 35):
            ws.row_dimensions[row].height = 23.3
        for row in range(36, 44):
            ws.row_dimensions[row].height = 23.3
        ws.row_dimensions[45].height = 23.3
        
        # Define styles
        bold_font = Font(name='Calibri', size=11, bold=True)
        bold_border = Border(
            left=Side(style='thick', color='000000'),
            right=Side(style='thick', color='000000'),
            top=Side(style='thick', color='000000'),
            bottom=Side(style='thick', color='000000')
        )
        
        # Section headers with bold outline
        ws['A5'] = 'Preliminary Architecture'
        ws['A5'].font = bold_font
        ws['A5'].border = bold_border
        
        ws['A28'] = 'Module Overview'
        ws['A28'].font = bold_font
        ws['A28'].border = bold_border
        
        # Column E should be unlined for specified sections (5-25, 28-43)
        # All other sections should have bold outline
        
        # Sheet ends at column F after row 66
        ws.sheet_properties.tabColor = "70AD47"
    
    def _create_item_definition_threat_model_sheet(self, workbook, analysis_data):
        """Create Item Definition - Threat Model sheet with precise formatting"""
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        ws = workbook.create_sheet('Item Definition - Threat Model')
        
        # Set precise column widths as specified
        ws.column_dimensions['B'].width = 137
        ws.column_dimensions['D'].width = 137
        
        # Define styles
        bold_font = Font(name='Calibri', size=11, bold=True)
        bold_border = Border(
            left=Side(style='thick', color='000000'),
            right=Side(style='thick', color='000000'),
            top=Side(style='thick', color='000000'),
            bottom=Side(style='thick', color='000000')
        )
        
        # Headers with bold outline and bold titles
        headers = ['ID', 'Threat Description', 'Category', 'Impact']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = bold_font
            cell.border = bold_border
        
        # Rest of sheet should be unlined except headers
        # Sheet ends at column E and after line 33
        ws.sheet_properties.tabColor = "FFC000"
    
    def _create_asset_identification_sheet(self, workbook, analysis_data):
        """Create Asset Identification sheet with precise formatting"""
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        ws = workbook.create_sheet('Asset Identification')
        
        # Set precise row heights as specified
        ws.row_dimensions[1].height = 25.5
        ws.row_dimensions[3].height = 21
        ws.row_dimensions[4].height = 29.1
        ws.row_dimensions[5].height = 14.7
        ws.row_dimensions[16].height = 14.7
        
        for row in range(6, 16):
            ws.row_dimensions[row].height = 14.4
        for row in range(17, 27):
            ws.row_dimensions[row].height = 14.4
        
        # Define styles
        bold_font = Font(name='Calibri', size=11, bold=True)
        bold_border = Border(
            left=Side(style='thick', color='000000'),
            right=Side(style='thick', color='000000'),
            top=Side(style='thick', color='000000'),
            bottom=Side(style='thick', color='000000')
        )
        
        # Title headers with bold outline and bold font
        ws['A1'] = 'Asset Identification'
        ws['A1'].font = bold_font
        ws['A1'].border = bold_border
        
        # Row 5 and 16 titles with bold outline
        ws['A5'] = 'Primary Assets'
        ws['A5'].font = bold_font
        ws['A5'].border = bold_border
        
        ws['A16'] = 'Supporting Assets'
        ws['A16'].font = bold_font
        ws['A16'].border = bold_border
        
        # Column headers and table with bold outline
        headers = ['Asset ID', 'Asset Name', 'Type', 'Description', 'Value', 'Owner', 'Location', 'Dependencies']
        for col_idx, header in enumerate(headers, start=1):
            for row in [6, 17]:  # Both table sections
                cell = ws.cell(row=row, column=col_idx, value=header)
                cell.font = bold_font
                cell.border = bold_border
        
        # Sheet ends at column I and after line 27
        ws.sheet_properties.tabColor = "E2EFDA"
    
    def _create_tara_sheet(self, workbook, analysis_data):
        """Create TARA sheet with precise formatting and frozen panes"""
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        ws = workbook.create_sheet('TARA')
        
        # Set precise row heights as specified
        ws.row_dimensions[1].height = 36
        ws.row_dimensions[3].height = 23.3
        ws.row_dimensions[4].height = 21.8
        
        # Odd rows 7-51 height 57.8 with background color
        for row in range(7, 52, 2):  # Odd rows only
            ws.row_dimensions[row].height = 57.8
        
        # Define styles
        bold_font = Font(name='Calibri', size=11, bold=True)
        bold_border = Border(
            left=Side(style='thick', color='000000'),
            right=Side(style='thick', color='000000'),
            top=Side(style='thick', color='000000'),
            bottom=Side(style='thick', color='000000')
        )
        
        # Color fills as specified
        threat_scenario_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        attack_feasibility_fill = PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid')
        cybersecurity_goals_fill = PatternFill(start_color='EDEDED', end_color='EDEDED', fill_type='solid')
        
        # Column headers as specified
        ws['B3'] = 'Asset ID'
        ws['C3'] = 'Asset'
        ws['D3'] = 'Security property loss'
        ws['E3'] = 'Stakeholder'
        
        # Apply bold font and border to main headers
        for col in ['B', 'C', 'D', 'E']:
            cell = ws[f'{col}3']
            cell.font = bold_font
            cell.border = bold_border
        
        # Threat scenario section (T-Y, line 3) with background color #FFF2CC
        ws.merge_cells('T3:Y3')
        ws['T3'] = 'Threat scenario'
        ws['T3'].font = bold_font
        ws['T3'].fill = threat_scenario_fill
        ws['T3'].border = bold_border
        ws['T3'].alignment = Alignment(horizontal='center')
        
        # Attack feasibility assessment (AA-AN, lines 3-5 and odd lines) with background #C6E0B4
        ws.merge_cells('AA3:AN3')
        ws['AA3'] = 'Attack feasibility assessment'
        ws['AA3'].font = bold_font
        ws['AA3'].fill = attack_feasibility_fill
        ws['AA3'].border = bold_border
        ws['AA3'].alignment = Alignment(horizontal='center')
        
        # Cybersecurity goals and claims (AO-AT, lines 3-5 and odd rows) with background #EDEDED
        ws.merge_cells('AO3:AT3')
        ws['AO3'] = 'Cybersecurity goals and claims'
        ws['AO3'].font = bold_font
        ws['AO3'].fill = cybersecurity_goals_fill
        ws['AO3'].border = bold_border
        ws['AO3'].alignment = Alignment(horizontal='center')
        
        # Apply header background colors to lines 3-5
        for row in range(3, 6):
            for col_idx in range(20, 25):  # T-Y columns
                cell = ws.cell(row=row, column=col_idx)
                cell.fill = threat_scenario_fill
                cell.border = bold_border
            
            for col_idx in range(27, 40):  # AA-AN columns
                cell = ws.cell(row=row, column=col_idx)
                cell.fill = attack_feasibility_fill
                cell.border = bold_border
            
            for col_idx in range(41, 46):  # AO-AT columns
                cell = ws.cell(row=row, column=col_idx)
                cell.fill = cybersecurity_goals_fill
                cell.border = bold_border
        
        # Apply background colors to odd rows 7-51 with same section colors
        for row in range(7, 52, 2):  # Odd rows
            for col_idx in range(20, 25):  # T-Y columns
                cell = ws.cell(row=row, column=col_idx)
                cell.fill = threat_scenario_fill
                cell.border = bold_border
            
            for col_idx in range(27, 40):  # AA-AN columns  
                cell = ws.cell(row=row, column=col_idx)
                cell.fill = attack_feasibility_fill
                cell.border = bold_border
            
            for col_idx in range(41, 46):  # AO-AT columns
                cell = ws.cell(row=row, column=col_idx)
                cell.fill = cybersecurity_goals_fill
                cell.border = bold_border
        
        # Bold black outline for headers, tables, and even rows
        for row in range(6, 53, 2):  # Even rows
            for col in range(2, 46):  # B to AT
                cell = ws.cell(row=row, column=col)
                cell.border = bold_border
        
        # Header should always be visible - freeze panes at lines 1-5
        ws.freeze_panes = 'A6'
        
        # Sheet ends at AU (column 47) and row 53
        ws.sheet_properties.tabColor = "FFD966"
    
    def _create_matrices_guidelines_sheet(self, workbook, analysis_data):
        """Create Matrices and Guidelines sheet with extremely precise formatting"""
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        ws = workbook.create_sheet('Matrices and Guidelines')
        
        # Define styles
        bold_font = Font(name='Calibri', size=11, bold=True, color='000000')
        normal_font = Font(name='Calibri', size=11, color='000000')
        bold_border = Border(
            left=Side(style='thick', color='000000'),
            right=Side(style='thick', color='000000'),
            top=Side(style='thick', color='000000'),
            bottom=Side(style='thick', color='000000')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        center_wrap_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Color fills as specified
        green_fill = PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')
        dark_green_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
        orange_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        
        # Column B rows 2-5 outlined in bold
        for row in range(2, 6):
            cell = ws.cell(row=row, column=2)
            cell.border = bold_border
        
        # Column D lines 5-8 center justified
        for row in range(5, 9):
            cell = ws.cell(row=row, column=4)
            cell.alignment = center_alignment
        
        # Columns K, M, O, Q, S center justified throughout
        center_cols = [11, 13, 15, 17, 19]  # K, M, O, Q, S
        for col in center_cols:
            for row in range(1, 30):
                cell = ws.cell(row=row, column=col)
                cell.alignment = center_alignment
        
        # Replace <= with ≤ symbol in column K anywhere it appears
        for row in range(1, 30):
            cell = ws.cell(row=row, column=11)  # Column K
            if cell.value and '<=' in str(cell.value):
                cell.value = str(cell.value).replace('<=', '≤')
        
        # Rows 4 and 5 columns V-AA with #C6E0B4 background and bold outline
        for row in [4, 5]:
            for col in range(22, 27):  # V-AA columns (22-26)
                cell = ws.cell(row=row, column=col)
                cell.fill = green_fill
                cell.border = bold_border
                if col in [22, 23]:  # V and W columns
                    cell.font = bold_font
        
        # Row 11 and 20 V-AA merged - "Threat analysis and risk assessment" with #FFD966 background
        for row in [11, 20]:
            ws.merge_cells(f'V{row}:AA{row}')
            cell = ws[f'V{row}']
            cell.value = 'Threat analysis and risk assessment'
            cell.font = bold_font
            cell.fill = yellow_fill
            cell.alignment = center_alignment
            cell.border = bold_border
        
        # Row 13 V-AA merged - "Risk threshold level" with #C6E0B4 background
        ws.merge_cells('V13:AA13')
        cell = ws['V13']
        cell.value = 'Risk threshold level'
        cell.font = bold_font
        cell.fill = green_fill
        cell.alignment = center_alignment
        cell.border = bold_border
        
        # V and W merged for rows 13, 14, 22, 23 with #C6E0B4 background
        merge_rows = [13, 14, 22, 23]
        for row in merge_rows:
            ws.merge_cells(f'V{row}:W{row}')
            cell = ws[f'V{row}']
            cell.fill = green_fill
            cell.font = bold_font
            cell.alignment = center_alignment
            cell.border = bold_border
        
        # X-AA line 13 merged - "Attack feasibility rating" with #C6E0B4 background
        ws.merge_cells('X13:AA13')
        cell = ws['X13']
        cell.value = 'Attack feasibility rating'
        cell.font = bold_font
        cell.fill = green_fill
        cell.alignment = center_alignment
        cell.border = bold_border
        
        # Line 14 attack feasibility labels with #C6E0B4 background
        feasibility_labels = [('X14', 'Very low'), ('Y14', 'Low'), ('Z14', 'Medium'), ('AA14', 'High')]
        for cell_ref, label in feasibility_labels:
            cell = ws[cell_ref]
            cell.value = label
            cell.font = bold_font
            cell.fill = green_fill
            cell.alignment = center_alignment
            cell.border = bold_border
        
        # Impact severity labels in column W with #C6E0B4 background
        severity_labels = [('W15', 'Severe'), ('W16', 'Major'), ('W17', 'Moderate'), ('W18', 'Negligible')]
        for cell_ref, label in severity_labels:
            cell = ws[cell_ref]
            cell.value = label
            cell.font = bold_font
            cell.fill = green_fill
            cell.alignment = center_alignment
            cell.border = bold_border
        
        # "Below" ratings with #70AD47 background
        below_cells = ['X15', 'X16', 'X17', 'X18', 'Y18', 'Z18', 'AA18']
        for cell_ref in below_cells:
            cell = ws[cell_ref]
            cell.value = 'Below'
            cell.fill = dark_green_fill
            cell.font = bold_font
            cell.alignment = center_alignment
            cell.border = bold_border
        
        # "Above" ratings with #FFC000 background
        above_cells = ['Y15', 'Y16', 'Y17', 'Z15', 'Z16', 'Z17', 'AA15', 'AA16', 'AA17']
        for cell_ref in above_cells:
            cell = ws[cell_ref]
            cell.value = 'Above'
            cell.fill = orange_fill
            cell.font = bold_font
            cell.alignment = center_alignment
            cell.border = bold_border
        
        # Row 21 V-AA merged - "CAL determination" with #C6E0B4 background
        ws.merge_cells('V21:AA21')
        cell = ws['V21']
        cell.value = 'CAL determination'
        cell.font = bold_font
        cell.fill = green_fill
        cell.alignment = center_alignment
        cell.border = bold_border
        
        # Row 22 X-AA merged - "Attack vector" with #C6E0B4 background  
        ws.merge_cells('X22:AA22')
        cell = ws['X22']
        cell.value = 'Attack vector'
        cell.font = bold_font
        cell.fill = green_fill
        cell.alignment = center_alignment
        cell.border = bold_border
        
        # Row 23 attack vector labels with #C6E0B4 background
        vector_labels = [('X23', 'Physical'), ('Y23', 'Local'), ('Z23', 'Adjacent'), ('AA23', 'Network')]
        for cell_ref, label in vector_labels:
            cell = ws[cell_ref]
            cell.value = label
            cell.font = bold_font
            cell.fill = green_fill
            cell.alignment = center_alignment
            cell.border = bold_border
        
        # V24-V27 merged - "Impact Rating" with #C6E0B4 background and text wrap
        ws.merge_cells('V24:V27')
        cell = ws['V24']
        cell.value = 'Impact Rating'
        cell.font = bold_font
        cell.fill = green_fill
        cell.alignment = center_wrap_alignment
        cell.border = bold_border
        
        # Impact severity labels in column W (24-27) with #C6E0B4 background
        impact_w_labels = [('W24', 'Severe'), ('W25', 'Major'), ('W26', 'Moderate'), ('W27', 'Negligible')]
        for cell_ref, label in impact_w_labels:
            cell = ws[cell_ref]
            cell.value = label
            cell.font = bold_font
            cell.fill = green_fill
            cell.alignment = center_alignment
            cell.border = bold_border
        
        # Impact rating matrix with specific values and colors as specified
        impact_values = {
            # Row 24: X=2(green), Y=3(orange), Z=4(red), AA=4(red)
            'X24': (2, dark_green_fill), 'Y24': (3, orange_fill), 'Z24': (4, red_fill), 'AA24': (4, red_fill),
            # Row 25: X=1(green), Y=2(green), Z=3(orange), AA=4(red)  
            'X25': (1, dark_green_fill), 'Y25': (2, dark_green_fill), 'Z25': (3, orange_fill), 'AA25': (4, red_fill),
            # Row 26: X=1(green), Y=1(green), Z=2(green), AA=3(orange)
            'X26': (1, dark_green_fill), 'Y26': (1, dark_green_fill), 'Z26': (2, dark_green_fill), 'AA26': (3, orange_fill)
        }
        
        for cell_ref, (value, fill) in impact_values.items():
            cell = ws[cell_ref]
            cell.value = value
            cell.fill = fill
            cell.font = bold_font
            cell.alignment = center_alignment
            cell.border = bold_border
        
        # Row 27 X-AA with "-" values
        for col_ref in ['X27', 'Y27', 'Z27', 'AA27']:
            cell = ws[col_ref]
            cell.value = '-'
            cell.font = bold_font
            cell.alignment = center_alignment
            cell.border = bold_border
        
        # Apply #C6E0B4 background to all specified ranges with black bold center justified text
        specified_ranges = [
            ('V13', 'AA14'),  # Lines 13 and 14
            ('V21', 'AA23'),  # Lines 21, 22, and 23
            ('V15', 'W18'),   # Column V and W lines 15-18
            ('V24', 'W27')    # Column V and W lines 24-27
        ]
        
        for start_cell, end_cell in specified_ranges:
            for row in ws[f'{start_cell}:{end_cell}']:
                for cell in row:
                    if cell.fill.start_color.rgb != dark_green_fill.start_color.rgb and \
                       cell.fill.start_color.rgb != orange_fill.start_color.rgb and \
                       cell.fill.start_color.rgb != red_fill.start_color.rgb and \
                       cell.fill.start_color.rgb != yellow_fill.start_color.rgb:
                        cell.fill = green_fill
                    cell.font = bold_font
                    cell.alignment = center_alignment
                    cell.border = bold_border
        
        ws.sheet_properties.tabColor = "C6E0B4"